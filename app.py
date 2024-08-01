#importações
from argparse import Action
from datetime import datetime
import time
import socket
import uuid
import subprocess
from PyPDF2 import PdfFileReader, PdfFileWriter
from flask import Flask, redirect, url_for, flash, request, render_template, session,jsonify, send_file, abort
from flask_mail import Mail, Message
from flask_sqlalchemy import SQLAlchemy
from decimal import Decimal, ROUND_HALF_UP
from fpdf import FPDF
import pandas as pd
import qrcode
from werkzeug.utils import secure_filename
from datetime import datetime
import os
from flask import send_from_directory
from flask import Flask, make_response
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from sqlalchemy import LargeBinary, func
from itsdangerous import URLSafeTimedSerializer
from flask import current_app
from itsdangerous import BadSignature, URLSafeTimedSerializer, SignatureExpired
from itsdangerous import URLSafeTimedSerializer, SignatureExpired,BadSignature
from flask import Flask, render_template, request, send_file,make_response
from openpyxl import Workbook
from io import BytesIO, StringIO
from werkzeug.security import generate_password_hash, check_password_hash
from flask_mail import Mail, Message
import pandas as pd
from sqlalchemy import func
from dashboards.dash_app import create_dash_app
from dashboards.dashA import create_dashboard_autorizador
from dashboards.dashBeneficiados import create_dashboard_beneficiado
from sqlalchemy import func
from datetime import datetime
from collections import Counter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
import requests
from routes.inicio import inicio_route
from routes.prefeitura import prefeitura_route
from routes.farmacia import farmacia_route
from routes.autorizador import autorizador_route 
from routes.monitoring import monitoring_bp, socketio
from models import db, Usuario, Farmacia, Transacao, TransacaoA, Autorizador, Prefeitura, Beneficiado, Dados, DadosA, UsuarioTemp,DadosTemporarios, PaginaInicial, Documento
from flask_admin import Admin,AdminIndexView,expose, BaseView, helpers as admin_helpers
from flask_admin.contrib.sqla import ModelView
from PIL import Image as PILImage
from routes.gerarcsv import gerarcsv_route
from flask_admin.actions import action
# import schedule
# import threading
# Criar instância do Flask
app = Flask(__name__, template_folder='templates')

#duração das flash message
app.config['MESSAGE_FLASHING_OPTIONS'] = {'duration': 5}
app.register_blueprint(inicio_route)
app.register_blueprint(prefeitura_route, url_prefix='/prefeitura')
app.register_blueprint(autorizador_route, url_prefix='/autorizador')
app.register_blueprint(farmacia_route, url_prefix='/farmacia', name='farmacia')
app.register_blueprint(gerarcsv_route)
app.register_blueprint(monitoring_bp, url_prefix='/monitoring') 

# Configurar o SQLAlchemy
app.config.from_pyfile('config.py')

# Configurar o envio de email
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587  # Porta do servidor SMTP do Gmail
app.config['MAIL_USE_TLS'] = True  # Usar TLS (Transport Layer Security)
app.config['MAIL_USERNAME'] = 'kevynlevi0@gmail.com'  # Seu email do Gmail
app.config['MAIL_PASSWORD'] = 'tzgw vnhu rbhp xftg'  # Sua senha do Gmail
app.config['MAIL_DEFAULT_SENDER'] = 'kevynlevi0@gmail.com'  # Configuração do remetente padrão
app.config['TIMEZONE'] = 'America/Sao_Paulo'

#tema flask admin
app.config['FLASK_ADMIN_SWATCH'] = 'cerulean'
s = URLSafeTimedSerializer(app.config['SECRET_KEY'],salt="activate")
mail = Mail(app)

# Configurar o LoginManager
login_manager = LoginManager(app)
login_manager.init_app(app)

# Definir as extensões de arquivo permitidas para upload
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}


def capitalize_string(s):
    return s.capitalize() if s else s

db.init_app(app)
# Models do banco de dados

class MyAdminIndexView(AdminIndexView):
    @expose('/')
    def index(self):
        if not current_user.is_authenticated or not current_user.is_admin:
            return redirect(url_for('login', next=request.url))
        
        if current_user.is_root:
            
            user_count = Usuario.query.count()
            active_user_count = Usuario.query.filter_by(is_active=True).count()
            admin_count = Usuario.query.filter_by(is_admin=True).count()
            root_count = Usuario.query.filter_by(is_root=True).count()

            return self.render(
                'adminDash.html', 
                user_count=user_count, 
                active_user_count=active_user_count, 
                admin_count=admin_count,
                root_count=root_count,
                tipo_usuario=current_user.tipo_usuario
            )
        else:
            
            user_count = Usuario.query.count()
            active_user_count = Usuario.query.filter_by(is_active=True).count()
            admin_count = Usuario.query.filter_by(is_admin=True).count()

            return self.render(
                'adminDash.html', 
                user_count=user_count, 
                active_user_count=active_user_count, 
                admin_count=admin_count,
                tipo_usuario=current_user.tipo_usuario
            )



class ServiceMonitorView(BaseView):
    @expose('/')
    def index(self):
        db_status = self.check_db_status()
        mail_status = self.check_mail_status()
        return self.render('service_monitor.html', db_status=db_status, mail_status=mail_status)
    
    def check_db_status(self):
        try:
            db.engine.execute('SELECT 1')
        except Exception as e:
            return "Offline"
        return "Online"
    
    def check_mail_status(self):
        try:
            mail.send(Message('Teste', recipients=[' ']))
        except Exception as e:
            return "Offline"
        return "Online"

    def is_accessible(self):
        return current_user.is_authenticated and current_user.is_admin

    def inaccessible_callback(self, name, **kwargs):
        return redirect(url_for('login', next=request.url))

class Controller(ModelView):
    can_delete = False
    column_list = ["email", "tipo_usuario", "is_active", "is_admin", "is_root"]
    
    @action(
        'ativar',
        'Ativar/Desativar Admin',
        'Você tem certeza que deseja alterar o status de admin dos usuários selecionados?'
    )
    def toggle_admin_status(self, ids):
        for user in Usuario.query.filter(Usuario.id.in_(ids)).all():
            if current_user.is_root:
                if not user.is_root:
                    user.is_admin = not user.is_admin
                    db.session.commit()
                    flash(f'Status administrativo de {user.email} foi alterado.', 'success')
                else:
                    flash(f'Você não pode alterar o status do usuário root {user.email}.', 'error')
            else:
                flash('Você não tem permissão para alterar status de administrador.', 'error')

    def is_accessible(self):
        return current_user.is_authenticated and (current_user.is_admin or current_user.is_root)

    def inaccessible_callback(self, name, **kwargs):
        flash(self.not_accessible_message(), 'error')
        return redirect(url_for('login', next=request.url))

    def not_accessible_message(self):
        return 'Faça login para acessar esta página'

    def on_model_change(self, form, model, is_created):
        if not current_user.is_root:
            if 'is_admin' in form or 'is_root' in form:
                flash('Você não tem permissão para alterar privilégios administrativos.', 'error')
                raise ValueError('Você não tem permissão para alterar privilégios administrativos.')
            
        if 'password' in form:
            model.password = generate_password_hash(form.password.data)

    def on_form_prefill(self, form, id):
        user = Usuario.query.get(id)
        if not current_user.is_root and user.is_root:
            flash('Você não tem permissão para editar o usuário root.', 'error')
            return redirect(url_for('admin.index'))
    def on_model_delete(self, model):
        if not current_user.is_root and model.is_root:
            flash('Você não tem permissão para excluir o usuário root.', 'error')
            raise ValueError('Você não tem permissão para excluir o usuário root.')

class LogView(BaseView):    
    @expose('/')
    def index(self):
        max_lines_per_page = 50  
        page = int(request.args.get('page', 1))
        start_line = (page - 1) * max_lines_per_page
        end_line = start_line + max_lines_per_page

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        try:
            with open('log.txt', 'r') as file:
                log_lines = file.readlines()

                
                if start_date and end_date:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d')
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                    log_lines = [line for line in log_lines if self.is_within_date_range(line, start_date, end_date)]

                total_lines = len(log_lines)
                log_content = log_lines[start_line:end_line]
        except Exception as e:
            log_content = [f"Erro ao ler o arquivo de log: {str(e)}"]
            total_lines = 0

        next_page = page + 1 if end_line < total_lines else None
        prev_page = page - 1 if start_line > 0 else None

        return self.render(
            'logs.html', 
            log_content=log_content, 
            next_page=next_page, 
            prev_page=prev_page, 
            start_date=request.args.get('start_date', ''), 
            end_date=request.args.get('end_date', '')
        )

    @expose('/download')
    def download_log(self):
        log_file_path = 'log.txt'

        # Geração do PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        try:
            with open(log_file_path, 'r') as file:
                for line in file:
                    pdf.cell(200, 10, txt=line.strip(), ln=True)
        except Exception as e:
            pdf.cell(200, 10, txt=f"Erro ao ler o arquivo de log: {str(e)}", ln=True)

        pdf_output_path = 'logfile.pdf'
        pdf.output(pdf_output_path)

        if os.path.exists(pdf_output_path):
            return send_file(   
                pdf_output_path,
                as_attachment=True,
                download_name='logfile.pdf',
                mimetype='application/pdf'
            )
        else:
            flash('Erro ao gerar o arquivo PDF.', 'error')
            return redirect(url_for('admin.index'))
    def is_within_date_range(self, line, start_date, end_date):
        try:
            log_date_str = line.split(' ')[0]  # Assume que a data está no início da linha
            log_date = datetime.strptime(log_date_str, '%Y-%m-%d')
            return start_date <= log_date <= end_date
        except Exception as e:
            return False

    def is_accessible(self):
        return current_user.is_authenticated and current_user.is_admin

    def inaccessible_callback(self, name, **kwargs):
        flash(self.not_accessible_message(), 'error')
        return redirect(url_for('login', next=request.url))

    def not_accessible_message(self):
        return 'Faça login para acessar esta página'

    
def security_context_processor():
    return dict(
        admin_base_template=admin.base_template,
        admin_view=admin.index_view,
        h=admin_helpers,
    )

admin = Admin(app, index_view=MyAdminIndexView(), template_mode='bootstrap4')

admin.add_view(Controller(Usuario, db.session))
admin.add_view(Controller(Beneficiado, db.session))
admin.add_view(ServiceMonitorView(name="Monitoramento de Serviços"))
admin.add_view(LogView(name="Logs do Sistema"))

# admin.add_view(Controller(Farmacia, db.session, name = "Farmácia"))

app.context_processor(security_context_processor)


def create_root_user():
    root_user = Usuario.query.filter_by(email='root@example.com').first()
    if not root_user:
        root_user = Usuario(
            email='root@example.com',
            password=generate_password_hash('rootpassword', method='sha256'),
            tipo_usuario='root',
            is_active=True,
            is_admin=True,
            is_root=True
        )
        db.session.add(root_user)
        db.session.commit()


#atualiza as informações do autorizador 
@app.route('/atualizarAutorizador/<cpf>', methods=['POST'])
def atualizarAutorizador(cpf):
    tipo_usuario = session.get('tipo_usuario','3')
    if tipo_usuario=='3':
        try:
            autorizador = Autorizador.query.filter_by(cpf=cpf).first()
            if not autorizador:
                return redirect(url_for('pagina_de_erro'))

            # Atualize as informações da farmácia com base nos dados do formulário enviado
            autorizador.nomeAutorizador = request.form['nomeAutorizador']
            autorizador.cpf = request.form['cpfAutorizador']


            # Commit para salvar as mudanças no banco de dados
            flash("Dados atualizados com sucesso!",'sucess')
            db.session.commit()

            # Redirecione de volta para a página de visualização da farmácia
            return '', 204
        except Exception as e:
            return render_template("prefeitura.html")
    else:
            print("entrou no else")
            return redirect(url_for('rota_protegida'))
from sqlalchemy.exc import SQLAlchemyError
import logging

@app.route('/atualizarBeneficiado/<cpf_beneficiado>', methods=['POST'])
@login_required
def atualizarBeneficiado(cpf_beneficiado):
    try:
        # Logs para depuração
        app.logger.info(f"Tentando atualizar o beneficiado com CPF: {cpf_beneficiado}")

        # Obter o beneficiado a partir do banco de dados utilizando Session.get()
        beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpf_beneficiado).first()
        if not beneficiado:
            flash("Beneficiado não encontrado.", "error")
            return redirect(url_for('login.html'))  # Redireciona para a página apropriada

        # Obtenha os dados do formulário
        nome_beneficiado = request.form.get('nome_beneficiado')
        cpf_beneficiado = request.form.get('cpf_beneficiado')
        cartao_sus = request.form.get('cartao_sus')
        nome_autorizado = request.form.get('nome_autorizado')
        cpf_autorizado = request.form.get('cpf_autorizado')
        quantidade_liberada = request.form.get('quantidade_liberada')
        tamanho_liberado = request.form.get('tamanho_liberado')
        motivo_liberacao = request.form.get('motivo_liberacao')
       
        validade_meses = request.form.get('validade_meses')

        # Validação de dados

        # Atualize as informações do beneficiado
        beneficiado.nome_beneficiado = nome_beneficiado
        beneficiado.cpf_beneficiado = cpf_beneficiado
        beneficiado.cartao_sus = cartao_sus
        beneficiado.nome_autorizado = nome_autorizado
        beneficiado.cpf_autorizado = cpf_autorizado
        beneficiado.quantidade_liberada = quantidade_liberada
        beneficiado.tamanho_liberado = tamanho_liberado
        beneficiado.motivo_liberacao = motivo_liberacao
      
        beneficiado.validade_meses = validade_meses

        # Commit para salvar as mudanças no banco de dados
        db.session.commit()
        flash("Informações do beneficiado atualizadas com sucesso!", "success")

        # Redirecione de volta para a página de visualização do beneficiado
        return '', 204

    except SQLAlchemyError as e:
        # Logging detalhado da exceção
        logging.error(f"Erro ao atualizar beneficiado: {e}", exc_info=True)
        flash("Erro ao atualizar beneficiado. Por favor, tente novamente mais tarde.", "error")
        return redirect(url_for('ativar_prefeitura'))  
    
@app.route('/atualizarBeneficiadoAT/<cpf_beneficiado>', methods=['POST'])
@login_required
def atualizarBeneficiadoAT(cpf_beneficiado):
    try:
        session['cpf_beneficiado'] = cpf_beneficiado
        
        # Logs para depuração
        app.logger.info(f"Tentando atualizar o beneficiado com CPF: {cpf_beneficiado}")

        # Obter o beneficiado a partir do banco de dados utilizando Session.get()
        beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpf_beneficiado).first()
        dadosTemporarios= DadosTemporarios
        
        if not beneficiado:
            flash("Beneficiado não encontrado.", "error")
            return redirect(url_for('ativar_prefeitura'))  # Redireciona para a página apropriada

        # Obtenha os dados do formulário

        quantidade_liberada = request.form.get('quantidade_liberada')
        tamanho_liberado = request.form.get('tamanho_liberado')
        motivo_liberacao = request.form.get('motivo_liberacao')
        confirmacao = request.form.get('confirmacao')
        tamanho_liberado=request.form.get('tamanho_liberado')
        
        beneficiado.confirmacao=confirmacao
        validade_meses = request.form.get('validade_meses')
        data_inicio = request.form.get('data_inicio')

        # Validação de dados
        print(confirmacao)
        print(type(confirmacao))
        confirmacao_bool = eval(confirmacao.capitalize())
        if confirmacao_bool==True:
            print("entra na confirmacao")
            dadostemporarios = DadosTemporarios(
                beneficiado_id=beneficiado.id,
                quantidade_liberada=quantidade_liberada,
                tamanho_liberado=tamanho_liberado,
                motivo_liberacao=motivo_liberacao,
                data_inicio=data_inicio,
                validade_meses=validade_meses
            )

            # Adicionando dados temporários ao session e commit
            db.session.add(dadostemporarios)
            beneficiado.confirmacao = True
            # Commit para salvar as mudanças no banco de dados
            db.session.commit()
            flash("Informações do beneficiado atualizadas com sucesso!", "success")

            # Redirecione de volta para a página de visualização do beneficiado
            return '', 204
        if confirmacao_bool==False:    
            # Atualize as informações do beneficiado
            beneficiado.quantidade_liberada = quantidade_liberada
            beneficiado.tamanho_liberado = tamanho_liberado
            beneficiado.motivo_liberacao = motivo_liberacao
            beneficiado.data_inicio= data_inicio
            beneficiado.validade_meses = validade_meses
            beneficiado.confirmacao = False
            # Commit para salvar as mudanças no banco de dados
            db.session.commit()
            flash("Informações do beneficiado atualizadas com sucesso!", "success")

            # Redirecione de volta para a página de visualização do beneficiado
            return '', 204
        else:
            return redirect(url_for('ativar_prefeitura'))  
    except SQLAlchemyError as e:
        # Logging detalhado da exceção
        logging.error(f"Erro ao atualizar beneficiado: {e}", exc_info=True)
        flash("Erro ao atualizar beneficiado. Por favor, tente novamente mais tarde.", "error")
        return redirect(url_for('ativar_prefeitura'))  
# Função para realizar a verificação e atualização dos dados
def verificacao():
    beneficiados = Beneficiado.query.all()  # Obtém todos os beneficiados do banco de dados
    
    for beneficiado in beneficiados:
        dados_temporarios = DadosTemporarios.query.filter_by(beneficiado_id=beneficiado.id).order_by(DadosTemporarios.id.desc()).first()
        
        if dados_temporarios:
            data_atual = datetime.now()
            
            if beneficiado.data_final < data_atual and beneficiado.confirmacao == 1:
                beneficiado.quantidade_liberada = dados_temporarios.quantidade_liberada
                beneficiado.tamanho_liberado = dados_temporarios.tamanho_liberado
                beneficiado.motivo_liberacao = dados_temporarios.motivo_liberacao
                beneficiado.validade_meses = dados_temporarios.validade_meses
                beneficiado.confirmacao=0
                data_final = beneficiado.data_inicio + relativedelta(months=beneficiado.validade_meses)
                data_final_formatada = data_final.strftime('%Y-%m-%d') 
                beneficiado.data_final= data_final_formatada
                beneficiado.data_inicio= datetime.now()
                try:
                    db.session.commit()
                    print(f'Dados atualizados para o beneficiado {beneficiado.id}')
                except Exception as e:
                    db.session.rollback()
                    print(f'Erro ao atualizar dados para o beneficiado {beneficiado.id}: {str(e)}')



# Método para gerar token de redefinição de senha
def get_reset_password_token(Usuario):
    serializer = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    token = serializer.dumps({'user_id': Usuario.id})
    return token

# Método para verificar token de redefinição de senha

def verify_reset_password_token(token, max_age=600):
    serializer = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    try:
        data = serializer.loads(token, max_age=max_age)
        usuario_id = data['user_id']  # Salva o ID do usuário do token
        return Usuario.Session.get(usuario_id)  # Busca o usuário pelo ID
    except:
        return None
    
def registar_log_alteracao_senha(usuario_id):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    
    data_alteracao= datetime.now().strftime('%Y-%m-%d %H:%M: %S')

    descricao_log =  f'''\nRegistro de alteração de senha: \n -Data:{data_alteracao}\n-id_usuario:{usuario_id}\n -Senha: {usuario_id.senha}\n - IP da Máquina: {ip_local}'''
    try:
      with open('log.txt','a') as arquivo_log:
        arquivo_log.write(descricao_log)
    except Exception as e:
     print(f"Erro ao escrever no arquivo de log: {e}")

# Rota para página de redefinição de senha
@app.route('/redef', methods=['GET', 'POST'])
def redef():
    try:
        if request.method == 'POST':
            email = request.form['username']
            usuario = Usuario.query.filter_by(email=email).first()
            if usuario:
                token = get_reset_password_token(usuario)
                send_password_reset_email(usuario, token)
                flash('Verifique seu e-mail para obter instruções sobre como redefinir sua senha.')
                return redirect(url_for('redef'))  # Redirecionar para a mesma página
            else:
                flash('E-mail não encontrado.')
        return render_template('redef.html', page = 'redef')
    except Exception as e:
        return redirect(url_for('redef'))

# Rota para página de redefinição de senha após clicar no link no e-mail
@app.route('/redefiDois/<token>', methods=['GET', 'POST'])
def redefiDois(token):
    try:
        usuario = verify_reset_password_token(token)
        if not usuario:
            flash('Token inválido ou expirado.')
            return redirect(url_for('redef'))  # Redirecionar para a página de redefinição de senha
        if request.method == 'POST':
            nova_senha = request.form['password']  # Supondo que você tenha um formulário para a nova senha
            usuario.senha = generate_password_hash(nova_senha)
            db.session.commit()
            flash('Sua senha foi redefinida com sucesso.')

            registar_log_alteracao_senha(usuario.id)

            return redirect(url_for('login'))  # Redirecionar para a página de login
        return render_template('redefiDois.html', token=token, page='redefiDois')
    except Exception as e:
        flash('Ocorreu um erro ao processar sua solicitação. Por favor, tente novamente mais tarde.')
        return redirect(url_for('redef'))  # Redirecionar para a página de redefinição de senha

# Função para enviar e-mail de redefinição de senha
def send_password_reset_email(user, token):
    reset_url = url_for('redefiDois', token=token, _external=True)
    msg = Message('Redefinição de Senha', sender='kevynlevi0@gmail.com', recipients=[user.email])
    msg.html = f'''<p>Você solicitou a redefinição de senha do sistema Easy Diaper, da Prefeitura Municipal de Vitória da Conquista. Por favor, <a href="{reset_url}">clique aqui</a> para redefinir sua senha.</p>
    <p>Se você não solicitou esta ação ou desconhece a tentativa, por favor ignore esta mensagem.</p>'''
    mail.send(msg)
#configurando o envio de email do fale conosco
@app.route('/enviar_email', methods=['POST'])
def enviar_email():
    if request.method == 'POST':
        email_remetente = request.form['email_remetente']
        
        assunto = request.form['assunto']
        mensagem = request.form['mensagem']
        app.config['MAIL_DEFAULT_SENDER'] = email_remetente
        msg = Message('Fale conosco-FRALDARIO', recipients=['kevynlevi0@gmail.com'])
        msg.body = f"De: {email_remetente}\n\n\n  Assunto: \name()   {assunto} \n\n \n \n Mensagem: \n\n \n \n{mensagem}"

        try:
            mail.send(msg)
            flash("email enviado com sucesso")
            return render_template("faleconosco.html")
        except Exception as e:
            print("Erro ao enviar o e-mail:", e)
            flash("erro ao enviar o email.")
            return render_template("faleconosco.html")
        
        



#configurando para gerar pdf da farmacia
# @app.route('/createPdf', methods=['POST'])
# def create_pdf():
#     # Obtém o CPF do beneficiado enviado do formulário
#     cpf_beneficiado = Beneficiado.cpfPesquisado
    
#     # Busca o beneficiado no banco de dados
#     beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpf_beneficiado).first()
    
#     if beneficiado:
#         # Cria um dicionário com os dados do beneficiado
#         data = {
#             'nome_beneficiado': beneficiado.nome_beneficiado,
#             'cpf_beneficiado': beneficiado.cpf_beneficiado,
#             'cartao_sus': beneficiado.cartao_sus,
#             'nome_autorizado': beneficiado.nome_autorizado,
#             'cpf_autorizado': beneficiado.cpf_autorizado,
#             'data_inicio': beneficiado.data_inicio.strftime("%d/%m/%Y"),  # Formata a data como string
#             'quantidade_pego': beneficiado.quantidade_pego,
#             'quantidade_restante': beneficiado.quantidade_restante,
#             'quantidade_liberada': beneficiado.quantidade_liberada,
#             'tamanho_liberado': beneficiado.tamanho_liberado,
#             'motivo_liberacao': beneficiado.motivo_liberacao,
#             'validade_meses': beneficiado.validade_meses,
#             'hora_atual': datetime.now().strftime("%H:%M"),  # Obtém a hora atual e formata como string
#         }
        
#         # Define o caminho do arquivo PDF
#         file_path = "documento.pdf"
        
#         # Chama a função para criar o PDF
#         create_pdf_file_Visu(file_path, data)
        
#         # Abre o arquivo PDF e envia como resposta
#         with open(file_path, 'rb') as f:
#             pdf_data = f.read()
        
#         response = make_response(pdf_data)
#         response.headers['Content-Type'] = 'application/pdf'
#         response.headers['Content-Disposition'] = 'inline; filename=documento.pdf'
        
        
#         return response
#     else:
#         return "Beneficiado não encontrado."
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.units import inch
import os

def create_pdf_file_Visu(file_path, data, protocol_number):
    # Cria o PDF
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    
    # Construa o texto contínuo
    text_lines = [
        f"Eu: {data.get('eu, ', '').strip()}",
        f"do CPF: {data.get('do cpf: ', '').strip()}",
        f"Estarei pegando: {data.get('Estarei pegando pegando ', '').strip()} quantidade de fraldas.",
        f"do beneficiado: {data.get('do beneficiado ', '').strip()}.",
        f"Cujo CPF: {data.get('Cujo CPF', '').strip()}.",
        f"No dia e hora: {data.get('No dia e hora', '').strip()}.",
        f"A data de início da autorização foi: {data.get('A data de início da autorização foi', '').strip()}.",
        f"Contando com a quantidade pega de hoje, restará: {data.get(', contando com a quantidade pega de hoje, restará ', '').strip()} fraldas.",
        f"Foi liberado: {data.get('Foi liberado ', '').strip()}.",
        f"Do tamanho: {data.get(', do tamanho ', '').strip()}.",
        f"Com a validade: {data.get('com a validade', '').strip()}."
    ]
    
    # Cria o texto contínuo
    text = ' '.join(text_lines).strip()  # Remove espaços extras e ajusta pontuação
    p = Paragraph(text, normal_style)
    
    # Adiciona o parágrafo ao PDF
    elements = [p]
    doc.build(elements)
@app.route('/createPdf', methods=['POST'])
def create_pdf():    
    # Obtém o CPF do beneficiado enviado do formulário
    cpf_beneficiado = Beneficiado.cpfPesquisado
    
    # Busca o beneficiado no banco de dados
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpf_beneficiado).first()
    
    if beneficiado: 
        # Cria um dicionário com os dados do beneficiado
        data = {
            'eu, ': beneficiado.nome_autorizado,
            'do cpf: ': beneficiado.cpf_autorizado,
            'Estarei pegando pegando ': beneficiado.quantidade_pego,
            'do beneficiado ': beneficiado.nome_beneficiado,
            'Cujo CPF': beneficiado.cpf_beneficiado,
            'No dia e hora': datetime.now().strftime("%d/%m/%Y, %H:%M:%S"),
            'A data de início da autorização foi': beneficiado.data_inicio.strftime("%d/%m/%Y"),
            ', contando com a quantidade pega de hoje, restará': beneficiado.quantidade_restante,
            'Foi liberado ': beneficiado.quantidade_liberada,
            ', do tamanho ': beneficiado.tamanho_liberado,
            'com a validade': beneficiado.validade_meses,
        }

        # Define o caminho do arquivo PDF
        protocol_number = uuid.uuid4().hex
        pdf_folder = "Pdfs_gerados_farmacia"
        os.makedirs(pdf_folder, exist_ok=True)
        file_path = os.path.join(pdf_folder, f"{protocol_number}.pdf")

        # Chama a função para criar o PDF
        create_pdf_file_Visu(file_path, data, protocol_number)
        
        # Abre o arquivo PDF e envia como resposta
        with open(file_path, 'rb') as f:
            pdf_data = f.read()
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=documento.pdf'
        
        return response
    else:
        return "Beneficiado não encontrado."


def generate_qr_code(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction = qrcode.constants.ERROR_CORRECT_L,
        box_size = 8,
        border =4,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    return img

def create_pdf_file_Visu(file_path, beneficiado_data,protocol_number):
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    qr_image = generate_qr_code(protocol_number)
    buffer = BytesIO()
    qr_image.save(buffer, format='PNG')
    buffer.seek(0)
    qr_code_image = Image(buffer, 1 *inch, 1 *inch)
    elements.append(qr_code_image)
    elements.append(Spacer(1, 12))
    
    # Adicionar título no cabeçalho
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # Centro
        spaceAfter=20,
    )
    title = Paragraph("Retirada de Fraldas", title_style)
    elements.append(title)

    protocol_style = ParagraphStyle(
        'Protocol',
        parent=styles['BodyText'],
        fontSize=12,
        spaceAfter=12,
    )
    protocol = Paragraph(f"Protocolo: {protocol_number}", protocol_style)
    elements.append(protocol)
    # Espaçamento após o título
    elements.append(Spacer(1, 12))
    
    # Estilo para os dados do beneficiado
    data_style = ParagraphStyle(
        'BodyText',
        parent=styles['BodyText'],
        fontSize=12,
        spaceAfter=12,
    )
    
    # Itera sobre cada chave e valor no dicionário de dados do beneficiado
    for key, value in beneficiado_data.items():
        # Formata a linha com a chave e o valor
        line = "{}: {}".format(key.replace('_', ' ').capitalize(), value)
        elements.append(Paragraph(line, data_style))
    
    # Espaçamento antes das assinaturas
    elements.append(Spacer(1, 40))
    
    # Adicionar linha separadora
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<hr width='100%'/>", styles["BodyText"]))
    elements.append(Spacer(1, 12))
    
    # Assinaturas
    elements.append(Paragraph("Assinatura do responsável:", data_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Assinatura do farmacêutico:", data_style))
    
    # Constrói o documento PDF
    doc.build(elements)

# Exemplo de uso da função
data = {
    'nome_beneficiado': 'Fulano de Tal',
    'cpf_beneficiado': '11111111111',
    'cartao_sus': '222222222222222',
    'nome_autorizado': 'Teste',
    'cpf_autorizado': '1212121212',
    'data_inicio': '17/06/2024',
    'quantidade_pego': '0',
    'quantidade_restante': '307',
    'quantidade_liberada': '1000',
    'tamanho_liberado': 'G',
    'motivo_liberacao': 'Sim',
    'validade_meses': '4',
    'hora_atual': '23:43'
}

protocol_number = uuid.uuid4().hex
create_pdf_file_Visu('documento_profissional.pdf', data, protocol_number)

def read_protocol_number(filename="protocol_number.txt"):
    if os.path.exists(filename):
        with open (filename, "r") as file:
            return int(file.read().strip())
    return 0

def save_protocol_number(number, filename="protocol_number.txt"):
    with open(filename, "w") as file:
        file.write(str(number))

@app.route('/createPdfFarmacia.html', methods=['POST'])
def createPdfFarmacia():
     # Consulta o banco de dados para obter informações sobre os beneficiados
    beneficiados = Beneficiado.query.all()
    
    # Calcula o total de pessoas que pegaram fraldas
    total_pessoas = len(beneficiados)
    
    # Calcula a quantidade total de fraldas retiradas e disponíveis
    total_retiradas = sum(b.quantidade_pego if b.quantidade_pego is not None else 0 for b in beneficiados)
    total_disponiveis = sum(b.quantidade_restante if b.quantidade_restante is not None else 0 for b in beneficiados)
    
    # Calcula a quantidade total de documentos anexados
    total_documentos = sum(1 for b in beneficiados if b.documento is not None or b.licitacao is not None)
    
    # Cria um dicionário com os dados do relatório
    relatorio_data = {
        'total_pessoas': total_pessoas,
        'total_retiradas': total_retiradas,
        'total_disponiveis': total_disponiveis,
        'total_documentos': total_documentos
    }
    
    # Define o caminho do arquivo PDF
    file_path = "documento.pdf"
    
    # Chama a função para criar o PDF
    create_pdf_file(file_path, beneficiados, relatorio_data)
    
    # Abre o arquivo PDF e envia como resposta
    with open(file_path, 'rb') as f:
        pdf_data = f.read()
    
    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename={protocol_number}.pdf'
    
    return response

def create_pdf_file(file_path, beneficiados, relatorio_data):
    c = canvas.Canvas(file_path, pagesize=letter)

    # Adiciona os dados do relatório de fraldas ao PDF
    relatorio_text = """
    <div class="relatorio">
        Relatório de Fraldas:
        Total de pessoas que pegaram fraldas: {total_pessoas}
        Quantidade total de fraldas retiradas: {total_retiradas}
        Quantidade total de fraldas disponíveis: {total_disponiveis}
    </div>
    """.format(**relatorio_data)

    # Define a posição inicial para o relatório
    y_position = 750

    # Desenha o texto do relatório
    for line in relatorio_text.split('\n'):
        c.drawString(100, y_position, line.strip())
        y_position -= 15  # Ajusta a posição vertical para a próxima linha

    # Adiciona os dados dos anexos dos beneficiados ao PDF
    anexos_text = """
    <div class="anexo">
        <h2>Anexos dos Beneficiados:</h2>
        <p>Total de Documentos: {total_documentos}</p>
    </div>
    """.format(**relatorio_data)

    # Define a posição inicial para os anexos
    y_position -= 30  # Espaço entre o relatório e os anexos
    for line in anexos_text.split('\n'):
        c.drawString(100, y_position, line.strip())
        y_position -= 15  # Ajusta a posição vertical para a próxima linha

    c.save()

# Verifica se o arquivo tem uma extensão permitida
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



import os
import re
import requests
import secrets
import string
from flask import Flask, request, render_template, redirect, flash, session, jsonify, url_for
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from datetime import datetime
from itsdangerous import URLSafeTimedSerializer
from flask_mail import Mail, Message
from PIL import Image as PILImage
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Image
from io import BytesIO
# Funções de validação
def is_valid_email(email: str) -> bool:
    email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(email_regex, email) is not None

def is_valid_cnpj(cnpj: str) -> bool:
    cnpj = ''.join(filter(str.isdigit, cnpj))
    if len(cnpj) != 14:
        return False

    def calculate_digit(cnpj, digit):
        if digit == 1:
            weights = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
        elif digit == 2:
            weights = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]

        sum_value = sum(int(cnpj[i]) * weights[i] for i in range(len(weights)))
        remainder = sum_value % 11
        return 0 if remainder < 2 else 11 - remainder

    calculated_digit1 = calculate_digit(cnpj, 1)
    calculated_digit2 = calculate_digit(cnpj, 2)

    return cnpj[-2:] == f'{calculated_digit1}{calculated_digit2}'

def is_valid_cpf(cpf: str) -> bool:
    cpf = ''.join(filter(str.isdigit, cpf))
    if len(cpf) != 11:
        return False
    return True

def generate_random_password(length=8):
    characters = string.ascii_letters + string.digits
    password = ''.join(secrets.choice(characters) for i in range(length))
    return password

def convert_image_to_pdf(image_path, output_path):
    image = PILImage.open(image_path)
    pdf = SimpleDocTemplate(output_path, pagesize=letter)
    elements = []
    buffer = BytesIO()
    image.save(buffer, format='PNG')
    buffer.seek(0)
    image_width, image_height = image.size
    aspect_ratio = image_height / float(image_width)
    image_pdf = Image(buffer)
    image_pdf.drawWidth = 6 * inch
    image_pdf.drawHeight = 6 * inch * aspect_ratio
    elements.append(image_pdf)
    pdf.build(elements)

def consultar_cnpj(cnpj):
    url = f'https://www.receitaws.com.br/v1/cnpj/{cnpj}'
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    else:
        return None

def consultar_cpf(cpf):
    # Você deve substituir isso com uma API real ou um serviço que você possua
    url = f'https://api.example.com/cpf/{cpf}'
    headers = {'Authorization': 'Bearer SEU_TOKEN_DE_ACESSO'}
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        return None
    
def limpar_cpf(cpf):
    # Remove pontos e traços do CPF
    cpf = cpf.replace('.', '').replace('-', '')
    return cpf
def limpar_cnpj(cnpj):
    # Remove pontos, traços e barra do CNPJ
    cnpj = cnpj.replace('.', '').replace('-', '').replace('/', '')
    return cnpj
def limpar_cep(cep):
    # Remove traços do CEP
    cep = cep.replace('-', '')
    return cep


@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    tipo_usuario = session.get('tipo_usuario', '3')
    nomePrefeitura = session.get('nomePrefeitura', '') if tipo_usuario == '3' else ''

    if request.method == 'POST':
        email = request.form['username']
        tipo_login = request.form.get('login')
        documento = request.files.get('documento')

        app.logger.info(f"Documento recebido: {documento}")
        if documento:
            app.logger.info(f"Nome do arquivo: {documento.filename}")
        else:
            app.logger.info("Nenhum documento recebido")

        if not is_valid_email(email):
            flash('E-mail inválido. Por favor, insira um e-mail válido.', 'error')
            return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura)

        if Usuario.query.filter_by(email=email).first() or UsuarioTemp.query.filter_by(email=email).first():
            flash('Este nome de usuário já está em uso. Por favor, escolha outro.', 'error')
            return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura)

        senha = generate_random_password()
        senha_hash = generate_password_hash(senha)

        cnpj = None
        nomeFantasia = None
        razaoSocial = None
        cep = None
        logradouro = None
        numero = None
        complemento = None
        bairro = None
        cidade = None
        estado = None
        tipos_fralda_str = ""
        tamanho_fralda_str = ""
        cpf = None
        nomeAutorizador = None
        cpf_prefeitura = None
        nomePrefeitura_input = None

        if not documento or documento.filename == '':
            flash('Por favor, envie o documento.', 'error')
            return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura)

        if not allowed_file(documento.filename):
            flash('Os arquivos devem ter uma extensão permitida.', 'error')
            return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura)

        filename = secure_filename(documento.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        documento.save(file_path)

        # Verificar e converter se necessário
        if filename.rsplit('.', 1)[1].lower() != 'pdf':
            pdf_filename = filename.rsplit('.', 1)[0] + '.pdf'
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
            convert_image_to_pdf(file_path, pdf_path)
            os.remove(file_path)  # Remove o arquivo de imagem original
            file_path = pdf_path

        with open(file_path, 'rb') as f:
            documento_blob = f.read()

        novo_usuario_temp = None

        try:
            if tipo_login == '1':  # Farmácia
                funcoes = request.form.get('funcoes', '') 
                cnpj = limpar_cnpj(request.form['cnpj'])
                nomeFantasia = request.form['nomeFantasia']
                razaoSocial = request.form['razaoSocial']
                cep = limpar_cep(request.form.get('cep', ''))
                logradouro = request.form.get('logradouro', '')
                numero = request.form.get('numero', '')
                complemento = request.form.get('complemento', '')
                bairro = request.form.get('bairro', '')
                cidade = request.form.get('cidade', '')
                estado = request.form.get('estado', '')
                tamanho_fralda = request.form.get('tamanho_fralda', '')
                tipos_fralda = request.form.get('tipos_fralda', '')

                if not is_valid_cnpj(cnpj):
                    flash('CNPJ inválido. Por favor, insira um CNPJ válido.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomeFantasia=nomeFantasia, cnpj=cnpj)

                if not logradouro or not numero or not cidade or not estado:
                    flash('Preencha todos os campos de endereço.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomeFantasia=nomeFantasia, cnpj=cnpj)

                # Verificar se o CNPJ já existe no banco de dados
                if Farmacia.query.filter_by(cnpj=cnpj).first():
                    flash('CNPJ já está em uso. Por favor, insira um CNPJ diferente.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomeFantasia=nomeFantasia, cnpj=cnpj)

                novo_usuario_temp = UsuarioTemp(
                    email=email, senha=senha_hash, tipo_login=tipo_login, cnpj=cnpj, nome=nomeFantasia, razaoSocial=razaoSocial, cep=cep,
                    logradouro=logradouro, numero=numero, complemento=complemento, bairro=bairro, cidade=cidade, estado=estado,
                    tipos_fralda=tipos_fralda, tamanho_fralda=tamanho_fralda, documento_liberacao=documento_blob, funcoes=funcoes
                )

            elif tipo_login == '2':  # Autorizador
                cpf = limpar_cpf(request.form['cpf'])
                nomeAutorizador = request.form['nomeAutorizador']
                funcoes = request.form.get('funcoes_autorizador', '')

                if not is_valid_cpf(cpf):
                    flash('CPF do autorizador inválido.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomeAutorizador=nomeAutorizador, cpf=cpf)

                if verificar_cpf_existente(cpf):
                    flash('CPF já está em uso.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomeAutorizador=nomeAutorizador, cpf=cpf)

                novo_usuario_temp = UsuarioTemp(
                    email=email, senha=senha_hash, tipo_login=tipo_login, cpf=cpf, nome=nomeAutorizador,
                    tipos_fralda="", tamanho_fralda="", documento_liberacao=documento_blob,funcoes=funcoes
                )

            elif tipo_login == '3':  # Prefeitura
                cpf_prefeitura = limpar_cpf(request.form.get('cpf_prefeitura', ''))
                nomePrefeitura_input = request.form.get('nomePrefeitura', '')
                funcoes = request.form.get('funcoes_prefeitura', '')

                if not is_valid_cpf(cpf_prefeitura):
                    flash('CPF da prefeitura inválido.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura_input, cpf_prefeitura=cpf_prefeitura)

                if verificar_cpf_existente(cpf_prefeitura):
                    flash('CPF já está em uso.', 'error')
                    return render_template("cadastro.html", username=email, tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura_input, cpf_prefeitura=cpf_prefeitura)

                novo_usuario_temp = UsuarioTemp(
                    email=email, senha=senha_hash, tipo_login=tipo_login, cpf=cpf_prefeitura, nome=nomePrefeitura_input,
                    tipos_fralda="", tamanho_fralda="", documento_liberacao=documento_blob,funcoes=funcoes
                )

            if novo_usuario_temp:
                app.logger.info(f"Tentando salvar o documento: {len(novo_usuario_temp.documento_liberacao)} bytes")
                db.session.add(novo_usuario_temp)
                db.session.commit()
                app.logger.info(f"Documento salvo: {novo_usuario_temp.documento_liberacao}")

                send_confirmation_email(email, senha)
                flash('Um email de confirmação foi enviado para você. Verifique seu email para confirmar seu cadastro.', 'success')

        except Exception as e:
            app.logger.error(f"Erro ao salvar o documento: {str(e)}")
            flash('Ocorreu um erro ao tentar salvar os dados. Por favor, tente novamente.', 'error')

    return render_template("cadastro.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)





@app.route('/consultar_cnpj/<cnpj>', methods=['GET'])
def consulta_cnpj_route(cnpj):
    dados_cnpj = consultar_cnpj(cnpj)
    if not dados_cnpj or 'status' in dados_cnpj and dados_cnpj['status'] == 'ERROR':
        return jsonify({'error': 'CNPJ inválido'}), 400
    return jsonify(dados_cnpj)



@app.route('/consultar_cpf/<cpf>', methods=['GET'])
def consulta_cpf_route(cpf):
    dados_cpf = consultar_cpf(cpf)
    if not dados_cpf:
        return jsonify({'error': 'CPF inválido'}), 400
    return jsonify(dados_cpf)

def buscar_endereco_cep(cep):
    try:
        response = requests.get(f'https://viacep.com.br/ws/{cep}/json')
        if response.status_code == 200:
            return response.json()
        else:
            return None
    except Exception as e:
        return None

def verificar_cpf_existente(cpf):
    return Autorizador.query.filter_by(cpf=cpf).first() or Prefeitura.query.filter_by(cpf_prefeitura=cpf).first() or Beneficiado.query.filter_by(cpf_beneficiado=cpf).first()

def send_confirmation_email(user_email, senha):
    token = s.dumps(user_email, salt='email-confirm')
    confirm_url = url_for('confirm_email', token=token, _external=True)
    html = render_template('email_template.html', confirm_url=confirm_url, senha=senha, user_email=user_email)
    msg = Message(subject='Confirme seu email', recipients=[user_email], html=html)
    mail.send(msg)


@app.route('/confirm_email/<token>', methods=['GET'])
def confirm_email(token):
    try:
        email = s.loads(token, salt='email-confirm')
        print(f"Token received: {token}")

        user_temp = UsuarioTemp.query.filter_by(email=email).first()
        print(f"Found user_temp: {user_temp}")
        
    except SignatureExpired:
        flash('O link de confirmação expirou.', 'error')
        return redirect(url_for('cadastro'))
    except BadSignature:
        flash('O link de confirmação é inválido.', 'error')
        return redirect(url_for('cadastro'))
    
    if user_temp is None:
        flash('Usuário não encontrado ou já confirmado.', 'error')
        return redirect(url_for('cadastro'))

    tipo_usuario_num = int(user_temp.tipo_login)

    print(f"Mapping tipo_login '{user_temp.tipo_login}' to tipo_usuario '{tipo_usuario_num}'")

    novo_usuario = Usuario(email=user_temp.email, senha=user_temp.senha, tipo_usuario=tipo_usuario_num, funcoes=user_temp.funcoes)
    novo_usuario.email_confirmed = True

    print(f"Creating new user with type: {user_temp.tipo_login}")

    db.session.add(novo_usuario)
    db.session.commit()

    if tipo_usuario_num == 1:
        cnpj = user_temp.cnpj
        nome = user_temp.nome
        razaoSocial = user_temp.razaoSocial
        cep = user_temp.cep
        logradouro = user_temp.logradouro
        numero = user_temp.numero
        complemento = user_temp.complemento
        bairro = user_temp.bairro
        cidade = user_temp.cidade
        estado = user_temp.estado
        tipos_fralda = user_temp.tipos_fralda
        tamanho_fralda = user_temp.tamanho_fralda
        documento_liberacao = user_temp.documento_liberacao

        nova_farmacia = Farmacia(id_usuario=novo_usuario.id, cnpj=cnpj, nomeFantasia=nome, razaoSocial=razaoSocial, cep=cep, logradouro=logradouro,
                                numero=numero, complemento=complemento, bairro=bairro, cidade=cidade, estado=estado,
                                tipos_fralda=tipos_fralda, documento_liberacao=documento_liberacao,
                                tamanho_fralda=tamanho_fralda)
        db.session.add(nova_farmacia)
        registrar_log_farmacia_confirmacao(nova_farmacia, novo_usuario.id)

    elif tipo_usuario_num == 2:
        cpf = user_temp.cpf
        nome = user_temp.nome
        documento_liberacao = user_temp.documento_liberacao

        novo_autorizador = Autorizador(id_usuario=novo_usuario.id, cpf=cpf, nomeAutorizador=nome, documento_liberacao=documento_liberacao)
        db.session.add(novo_autorizador)
        registrar_log_autorizador_confirmacao(novo_autorizador, novo_usuario.id)

    elif tipo_usuario_num == 3:
        cpf = user_temp.cpf
        nome = user_temp.nome
  
        nova_prefeitura = Prefeitura(id_usuario=novo_usuario.id, cpf_prefeitura=cpf, nomePrefeitura=nome)
        db.session.add(nova_prefeitura)
        registrar_log_prefeitura_confirmado(nova_prefeitura, novo_usuario.id)

    db.session.delete(user_temp)
    db.session.commit()
    
    
    flash('Email confirmado com sucesso!', 'success')
    return redirect(url_for('login'))


import logging

from io import BytesIO
from sqlalchemy import func

# Configurar o logging se não estiver configurado
logging.basicConfig(level=logging.DEBUG)
from datetime import datetime
from dateutil.relativedelta import relativedelta
from flask import request, render_template, flash
from sqlalchemy.sql import func, extract
import pytz

@app.route('/cadastro_beneficiado', methods=['POST'])
def cadastro_beneficiado():
    tipo_usuario = session.get('tipo_usuario', '2')
    nomeAutorizador = session.get('nomeAutorizador', '') if tipo_usuario == '2' else ''

    if request.method == 'POST':
        try:
            current_month = datetime.now().month
            current_year = datetime.now().year

            id_autorizador = current_user.id if current_user.is_authenticated else None
            autorizador = Autorizador.query.filter_by(id_usuario=id_autorizador).first()

            if not autorizador:
                flash('Usuário autorizador não encontrado.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            nome_beneficiado = request.form['username']
            cpf_beneficiado = request.form['cpfBeneficiado']
            cartao_sus = request.form['cartaoSus']
            nome_autorizado = request.form['pessoaAutorizada']
            cpf_autorizado = request.form['cpfPessoa']
            quantidade_liberada = request.form['qntddLiberada']
            tamanho_liberado = request.form['tamanhoLiberada']
            motivo_liberacao = request.form['motivoLiberacao']
            marca_fralda = request.form['marca']
            data_inicio_str = request.form['dataInicio']  # data_inicio como string
            validade_meses = int(request.form['vigencia'])  # Convertendo validade_meses para int
            documento = request.files['documento']

            if not is_valid_cpf(cpf_beneficiado):
                flash('CPF do beneficiado inválido.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            if not is_valid_cpf(cpf_autorizado):
                flash('CPF do autorizado inválido.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            if Beneficiado.query.filter_by(cpf_beneficiado=cpf_beneficiado).first():
                flash('O CPF do beneficiado já está cadastrado.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            if not documento or not documento.filename:
                flash('Por favor, envie o documento.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            if not allowed_file(documento.filename):
                flash('Os arquivos devem ter uma extensão permitida.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            try:
                # Converte a string para um objeto datetime
                data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
                data_inicio_formatada = data_inicio.strftime('%Y-%m-%d')  # Formato para o banco de dados

                # Calcula a data final
                data_final = data_inicio + relativedelta(months=validade_meses)
                data_final_formatada = data_final.strftime('%Y-%m-%d')  # Formato para o banco de dados
                
            except ValueError:
                flash('Data de início inválida. Use o formato AAAA-MM-DD.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            documento_blob = documento.read()
            if len(documento_blob) > 65535:
                flash('O arquivo é muito grande. O tamanho máximo permitido é 64 KB.', 'error')
                return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

            # Cria um novo beneficiado
            novo_beneficiado = Beneficiado(
                ativo=1,
                data_final=data_final_formatada,
                nome_beneficiado=nome_beneficiado,
                cpf_beneficiado=cpf_beneficiado,
                cartao_sus=cartao_sus,
                nome_autorizado=nome_autorizado,
                cpf_autorizado=cpf_autorizado,
                quantidade_liberada=quantidade_liberada,
                tamanho_liberado=tamanho_liberado,
                motivo_liberacao=motivo_liberacao,
                marca_fralda=marca_fralda,
                data_inicio=data_inicio,
                validade_meses=validade_meses,
                documento=documento_blob,
                quantidade_restante=quantidade_liberada,
                id_autorizador=id_autorizador,
                data_cadastro=datetime.now(pytz.UTC),  # Data e hora atual em UTC
            )

            db.session.add(novo_beneficiado)
            db.session.commit()

            # Calcula os totais
# Consulta SQLAlchmey para obter o total de beneficiados por mês
            total_por_mes = TransacaoA.query.filter(
                extract('month', TransacaoA.data_retirada) == current_month,
                extract('year', TransacaoA.data_retirada) == current_year,
                TransacaoA.autorizador_id == autorizador.id
            ).with_entities(func.sum(TransacaoA.quantidadeBeneficiado)).scalar()

            # Se total_por_mes for None (ou seja, nenhum resultado retornado), definir como 0
            total_por_mes = total_por_mes or 0

            tz = pytz.timezone('America/Sao_Paulo')
            hora_atual = datetime.now(tz)
            nova_transacao = TransacaoA(
                autorizador_id=autorizador.id,
                detalhes_beneficiados=nome_beneficiado,
                total_por_mes=total_por_mes,
                cpf_beneficiadoA=cpf_beneficiado,
                data_retirada=hora_atual
            )

            db.session.add(nova_transacao)
            db.session.commit()

            registrar_log_beneficiado(novo_beneficiado, current_user.id if current_user.is_authenticated else 'Usuário Anônimo')

            flash('Cadastro realizado com sucesso!', 'success')
            return render_template("cadastroBenAut.html", total_por_mes=total_por_mes, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

        except Exception as e:
            app.logger.error(f"Erro durante o cadastro de beneficiado: {str(e)}")
            flash("Ocorreu um erro durante o cadastro. Por favor, tente novamente mais tarde.", 'error')
            return render_template("cadastroBenAut.html", form_data=request.form, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

    return render_template("cadastroBenAut.html", nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

@app.route('/cadastro_Pbeneficiado', methods=['GET', 'POST'])
def cadastro_Pbeneficiado():
    tipo_usuario = session.get('tipo_usuario', '3')
    nomePrefeitura = session.get('nomePrefeitura', '') if tipo_usuario == '3' else ''

    if request.method == 'POST':
        try:
            current_month = datetime.now().month
            current_year = datetime.now().year

            id_usuario = current_user.id if current_user.is_authenticated else None
            prefeitura = Prefeitura.query.filter_by(id_usuario=id_usuario).first()

            if not prefeitura:
                flash('Usuário prefeitura não encontrado.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            nome_beneficiado = request.form['username']
            cpf_beneficiado = request.form['cpfBeneficiado']
            cartao_sus = request.form['cartaoSus']
            nome_autorizado = request.form['pessoaAutorizada']
            cpf_autorizado = request.form['cpfPessoa']
            quantidade_liberada = request.form['qntddLiberada']
            tamanho_liberado = request.form['tamanhoLiberada']
            motivo_liberacao = request.form['motivoLiberacao']
            marca_fralda = request.form['marca']
            data_inicio = request.form['dataInicio']
            validade_meses = request.form['vigencia']
            documento = request.files['documento']

            if not is_valid_cpf(cpf_beneficiado):
                flash('CPF do beneficiado inválido.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            if not is_valid_cpf(cpf_autorizado):
                flash('CPF do autorizado inválido.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            if Beneficiado.query.filter_by(cpf_beneficiado=cpf_beneficiado).first():
                flash('O CPF do beneficiado já está cadastrado.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            if not documento or not documento.filename:
                flash('Por favor, envie o documento.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            if not allowed_file(documento.filename):
                flash('Os arquivos devem ter uma extensão permitida.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            documento_blob = documento.read()
            if len(documento_blob) > 65535:
                flash('O arquivo é muito grande. O tamanho máximo permitido é 64 KB.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            try:
                data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            except ValueError:
                flash('Data de início inválida. Use o formato AAAA-MM-DD.', 'error')
                return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            novo_beneficiado = Beneficiado(
                nome_beneficiado=nome_beneficiado,
                cpf_beneficiado=cpf_beneficiado,
                cartao_sus=cartao_sus,
                nome_autorizado=nome_autorizado,
                cpf_autorizado=cpf_autorizado,
                quantidade_liberada=quantidade_liberada,
                tamanho_liberado=tamanho_liberado,
                motivo_liberacao=motivo_liberacao,
                marca_fralda=marca_fralda,
                data_inicio=data_inicio_dt,
                validade_meses=validade_meses,
                documento=documento_blob,
                quantidade_restante=quantidade_liberada,
                id_prefeitura=prefeitura.id
            )

            db.session.add(novo_beneficiado)
            db.session.commit()

            registrar_log_beneficiado(novo_beneficiado, current_user.id if current_user.is_authenticated else 'Usuário Anônimo')

            flash('Cadastro realizado com sucesso!', 'success')
            return render_template("cadastroPbeneficiado.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

        except Exception as e:
            app.logger.error(f"Erro durante o cadastro de beneficiado: {str(e)}")
            flash("Ocorreu um erro durante o cadastro. Por favor, tente novamente mais tarde.", 'error')
            return render_template("cadastroPbeneficiado.html", form_data=request.form, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

    return render_template("cadastroPbeneficiado.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)


#rota de configuração de rotas pendentes
@app.route("/mostrarPendencia", methods=['POST', 'GET'])
@login_required
def mostrar_pendencia():
        return render_template("uPendenteF.html")


from flask import Flask, request, redirect, render_template, flash, session, send_file, make_response, jsonify, url_for
from sqlalchemy.exc import SQLAlchemyError

from io import BytesIO
from PyPDF2 import PdfMerger, PdfFileReader
from PIL import Image
import os
from werkzeug.utils import secure_filename
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

# Função para verificar extensões permitidas de arquivo
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Função para salvar arquivo temporário
def save_temporary_file(file):
    temp_file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(temp_file_path)
    return temp_file_path

# Função para mesclar documentos PDF e converter imagens em PDF antes de mesclar
def merge_documents(beneficiado):
    merger = PdfMerger()

    for documento in beneficiado.documentos:
        try:
            if documento.licitacao.startswith(b'%PDF'):
                merger.append(BytesIO(documento.licitacao))
            elif documento.licitacao.startswith(b'\x89PNG\r\n\x1a\n'):
                img = Image.open(BytesIO(documento.licitacao))
                pdf_buffer = BytesIO()
                img.save(pdf_buffer, format='PDF')
                pdf_buffer.seek(0)
                merger.append(pdf_buffer)
            else:
                flash(f'O documento de tipo não suportado foi ignorado.', 'error')
        except Exception as e:
            flash(f'Erro ao processar documento: {str(e)}', 'error')
            app.logger.error(f'Erro ao processar documento: {str(e)}')

    merged_pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'documentos_combinados.pdf')
    merger.write(merged_pdf_path)
    merger.close()

    return merged_pdf_path

# Rota para salvar documentos
@app.route("/salvaDocumento", methods=['POST'])
def salva_documento():
    tipo_usuario = session.get('tipo_usuario')
    if 'nomeFantasia' in session:
        nomeFantasia = session['nomeFantasia']

    cpfPesquisado = session.get('cpf_pesquisado')
    app.logger.info(f"CPF pesquisado: {cpfPesquisado}")
    if not cpfPesquisado:
        flash("CPF não configurado na sessão.")
        return redirect('/pagina_prefeitura')

    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
    app.logger.info(f"Beneficiado encontrado: {beneficiado}")

    if beneficiado is None:
        flash("Beneficiado não encontrado para o CPF fornecido.")
        return redirect('/some_route')

    if 'documentoBeneficiado' not in request.files:
        flash("Envio pendente: nenhum documento enviado.")
        return render_template('geraDocumento.html', beneficiado=beneficiado, cpfPesquisado=cpfPesquisado, pendente=True)

    files = request.files.getlist('documentoBeneficiado')

    if not files:
        flash("Envio pendente: nenhum documento enviado.")
        return render_template('geraDocumento.html', beneficiado=beneficiado, cpfPesquisado=cpfPesquisado, pendente=True)

    temp_files = []

    try:
        for file in files:
            if file.filename == '':
                continue

            if not allowed_file(file.filename):
                flash(f'O arquivo {file.filename} não tem uma extensão permitida.', 'error')
                continue

            temp_file_path = save_temporary_file(file)
            temp_files.append(temp_file_path)
            app.logger.info(f"Arquivo temporário salvo: {temp_file_path}")

            with open(temp_file_path, 'rb') as f:
                file_content = f.read()
                novo_documento = Documento(licitacao=file_content)
                beneficiado.documentos.append(novo_documento)

        db.session.commit()
        app.logger.info("Arquivos salvos com sucesso no banco de dados.")

        merged_pdf_path = merge_documents(beneficiado)
        app.logger.info(f"Documento PDF combinado salvo em: {merged_pdf_path}")

        with open(merged_pdf_path, 'rb') as f:
            combined_pdf_content = f.read()
            documento_combinado = Documento(licitacao=combined_pdf_content)
            beneficiado.documentos.append(documento_combinado)

        db.session.commit()
        app.logger.info("Documento combinado salvo no banco de dados.")

        return make_response('', 204)

    except SQLAlchemyError as e:
        flash(f'Erro de banco de dados: {str(e)}', 'error')
        db.session.rollback()
        app.logger.error(f'Erro de banco de dados ao salvar documento: {str(e)}')
        return redirect('/pagina_erro')

    except Exception as e:
        flash(f'Erro ao processar documentos: {str(e)}', 'error')
        app.logger.error(f'Erro ao processar documentos: {str(e)}')
        return redirect('/pagina_erro')

    finally:
        for file_path in temp_files:
            try:
                os.remove(file_path)
                app.logger.info(f'Arquivo temporário removido: {file_path}')
            except Exception as e:
                flash(f'Erro ao remover arquivo temporário {file_path}: {str(e)}', 'error')
                app.logger.error(f'Erro ao remover arquivo temporário {file_path}: {str(e)}')

    return make_response('', 204)

@app.route('/get-file')
def get_file():
    cpfPesquisado = Beneficiado.cpfPesquisado
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
    if beneficiado and beneficiado.documento:
        documento_blob = beneficiado.documento
        file_name = f"{cpfPesquisado}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        return send_file(
            io.BytesIO(documento_blob),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=file_name  # Usando download_name para definir o nome do arquivo
        )
    else:
        return "Arquivo não encontrado", 404



from flask import jsonify

@app.route('/list-files')
def list_files():
    cpfPesquisado = Beneficiado.cpfPesquisado
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
    if beneficiado:
        documento_blob = beneficiado.documento
        if documento_blob:
            # Aqui você pode retornar uma lista de informações de arquivo
            return jsonify([{
                'filename': f"{cpfPesquisado}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf",
                'url': url_for('get_file', cpfPesquisado=cpfPesquisado)
            }])
    return jsonify([])  # Retorna uma lista vazia se não houver arquivos
# Rota para obter o documento da prefeitura

# Rota para obter um documento da prefeitura
@app.route('/get-documento-prefeitura')
def get_documento_prefeitura():
    cpfPesquisado = session.get('cpf_pesquisado')
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()

    if beneficiado and beneficiado.documentos:
        documento_blob = beneficiado.documentos[-1].licitacao  # Supõe que o último documento é o combinado
        file_name = f"{cpfPesquisado}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        return send_file(
            BytesIO(documento_blob),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=file_name
        )
    else:
        return "Arquivo não encontrado", 404

@app.route('/get-licitacao-prefeitura')
def get_licitacao_prefeitura():
    cpfPesquisado = Beneficiado.cpfPesquisado
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
    if beneficiado and beneficiado.licitacao:
        documento_blob = beneficiado.licitacao
        file_name = f"{cpfPesquisado}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        return send_file(
            io.BytesIO(documento_blob),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=file_name 
        )
    else:
        return "Arquivo não encontrado", 404

# Rota para listar documentos da prefeitura
@app.route('/list-files-prefeitura')
def list_files_prefeitura():
    cpfPesquisado = session.get('cpf_pesquisado')
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()

    if beneficiado:
        documentos = []

        if beneficiado.documentos:
            documentos.append({
                'filename': f"{cpfPesquisado}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf",
                'url': url_for('get_documento_prefeitura')
            })

        return jsonify(documentos)

    return jsonify([])

@app.route('/pesquisar', methods=['POST'])
def pesquisar_beneficiado():
    if request.method == 'POST':
        cpfPesquisado = request.form['cpf_pesquisado']
        beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
        
        registro_de_pesquisa(cpfPesquisado, beneficiado is not None)
        
        if beneficiado:
            # Verifica se os anexos estão presentes
            if not (beneficiado.documento and beneficiado.licitacao):
                session['pendente'] = True  # Define a variável de sessão para indicar que o usuário está pendente
                
                return redirect(url_for('usuarioPendente'))
            else:
                # Se os anexos estiverem presentes, redireciona para a página de visualização do beneficiado
                return redirect(url_for('visualizaBeneficiadoUm'))
        else:
            flash('Beneficiado não encontrado.', 'error')
            return redirect(url_for('visualizaBenef'))
        
def registro_de_pesquisa(cpfPesquisado, beneficiado_encontrado):
    data_pesquisa = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    status = 'encontrado' if beneficiado_encontrado else 'não encontrado'
    descricao_log = f'Pesquisa realizada:\n- Data: {data_pesquisa}\n- CPF Pesquisado: {cpfPesquisado}\n- Status: {status}\n'
    
    try:
        with open('log.txt', 'a') as arquivo_log:
            arquivo_log.write(descricao_log)
    except Exception as e:
        print(f"Erro ao escrever no arquivo de log: {e}")
@app.route('/pesquisarFarmacia', methods=['POST'])
def pesquisar_farmacia():
    nomeAutorizador = ""
    nomeFantasia = ""
    tipo_usuario = session.get('tipo_usuario')

    try:
        if 'nomeAutorizador' in session:
            nomeAutorizador = session['nomeAutorizador']
        elif 'nomeFantasia' in session: 
            nomeFantasia = session ['nomeFantasia']

        if request.method == 'POST':
            cpf_pesquisado = request.form.get('cpf_pesquisado')
            session['cpf_pesquisado'] = cpf_pesquisado
            Beneficiado.cpfPesquisado = cpf_pesquisado

            beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=Beneficiado.cpfPesquisado).first()

            if beneficiado:
                data_final = beneficiado.data_final
                data_final_formatada = data_final.strftime('%d/%m/%Y')

                db.session.commit()  # Salvar a atualização no banco de dados

                # Impedir repetição
                impedir_repeticao = 1
                session['impedir_repeticao'] = impedir_repeticao

                # Renderizar o template visu.html e passar os dados do beneficiado como argumentos
                return render_template("visu.html", page='visuF', beneficiado=beneficiado, data_final_formatada=data_final_formatada, cpfPesquisado=Beneficiado.cpfPesquisado, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario, nomeFantasia=nomeFantasia)
            else:
                flash('Beneficiado não encontrado.', 'error')
                return render_template('busca.html', nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario, nomeFantasia=nomeFantasia)

        else:
            flash('Método HTTP inválido para esta rota.', 'error')
            return render_template('farmacia.html')

    except SQLAlchemyError as e:
        flash(f'Erro de banco de dados: {str(e)}', 'error')
        db.session.rollback()
        return redirect('portal')

    except Exception as e:
        flash(f'Erro inesperado: {str(e)}', 'error')
        return redirect('portal')
    
@app.route('/pesquisarAbeneficiado', methods=['POST'])
@login_required
def pesquisarAbeneficiado():
    tipo_usuario = session['tipo_usuario']
    if 'nomeAutorizador' in session:
        nomeAutorizador = session['nomeAutorizador']
    if request.method == 'POST':
        cpf_pesquisado = request.form.get('cpf_pesquisado', '').strip()
        
        # Debugging statement
        print(f"CPF Pesquisado: {cpf_pesquisado}")

        if  cpf_pesquisado == "":
            return render_template('busca.html', tipo_usuario=tipo_usuario)
            
            
        
        Beneficiado.cpfPesquisado = cpf_pesquisado
        beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpf_pesquisado).first()

        registro_de_pesquisa(cpf_pesquisado, beneficiado is not None)

        if beneficiado:
            db.session.commit()  # Save the update to the database
            return render_template("visu.html", page='visuA', beneficiado=beneficiado, cpfPesquisado=cpf_pesquisado, tipo_usuario=tipo_usuario, nomeAutorizador=nomeAutorizador)
        else:
            flash('Beneficiado não encontrado.', 'error')
            return render_template('busca.html')
    else:
        flash('Método HTTP inválido para esta rota.', 'error')
        return render_template('autorizador.html', tipo_usuario=tipo_usuario, nomeAutorizador=nomeAutorizador)
  # Redireciona para a página da farmácia
def registro_documentacao_gerada(cpfPesquisado, quantidade_pego, operacao_bem_sucedida):
    data_operacao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    status = 'sucesso' if operacao_bem_sucedida else 'falha'
    descricao_log = f'Geração de documento:\n- Data: {data_operacao}\n- CPF Pesquisado: {cpfPesquisado}\n- Quantidade Pegada: {quantidade_pego}\n- Status: {status}\n'
    
    try:
        with open('log_documentacao_gerada.txt', 'a') as arquivo_log:
            arquivo_log.write(descricao_log)
    except Exception as e:
        print(f"Erro ao escrever no arquivo de log: {e}")
@app.route('/gerar_dados_mensais')
def gerar_dados_mensais():
    # Obter o mês e ano atual
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year

    # Calcular médias mensais
    dados_mensais = db.session.query(
        func.avg(Farmacia.quantidade).label('media_quantidade_fralda'),
        func.avg(Farmacia.tamanho_fralda).label('media_tamanho_fralda')
    ).filter(
        extract('month', Farmacia.data_registro) == mes_atual,
        extract('year', Farmacia.data_registro) == ano_atual
    ).first()

    # Se não houver dados para o mês atual, retorna zero para evitar erros
    if not dados_mensais.media_quantidade_fralda:
        dados_mensais.media_quantidade_fralda = 0

    if not dados_mensais.media_tamanho_fralda:
        dados_mensais.media_tamanho_fralda = 0

    # Atualiza ou insere dados mensais na sua lógica de persistência
    # Aqui, vamos imprimir os dados para fins de demonstração
    print(f'Média Quantidade Fralda: {dados_mensais.media_quantidade_fralda}')
    print(f'Média Tamanho Fralda: {dados_mensais.media_tamanho_fralda}')

    # Poderia ser inserido ou atualizado na tabela DadosMensais ou outro local de persistência

    return render_template('gerar_dados_mensais.html', dados_mensais=dados_mensais)

from datetime import datetime, timedelta, timezone
from apscheduler.schedulers.background import BackgroundScheduler 
from flask import current_app

# Define o intervalo de tempo para processar os dados ao final do mês
PROCESSAMENTO_FINAL_MES = timedelta(days=1)  # Um dia após o final do mês

# Função para processar os dados ao final do mês
def processar_dados_ao_final_do_mes():
    # Busca todas as transações temporárias até a data atual
    transacoes_temporarias = Transacao.query.filter(Transacao.data_retirada < datetime.utcnow()).all()
    
    for transacao in transacoes_temporarias:
        # Define a data de armazenamento permanente como o primeiro dia do próximo mês
        transacao.data_permanente = datetime.utcnow().replace(day=1) + timedelta(days=32)
    
    db.session.commit()

    # Limpa todas as transações temporárias após o processamento
    Transacao.query.filter(Transacao.data_retirada < datetime.utcnow()).delete()
    db.session.commit()

# Agendamento para executar a função ao final do mês
scheduler = BackgroundScheduler()
scheduler.add_job(processar_dados_ao_final_do_mes, 'cron', day='1', hour='0', minute='0')  # Executar no primeiro dia do mês às 00:00
scheduler.start()
scheduler = BackgroundScheduler()
scheduler.add_job(verificacao, 'cron', day='*', hour='0', minute='0')
scheduler.start()
from collections import defaultdict


def sugerir_media_mensal_marca():
    agora = datetime.now()
    mes_atual = agora.month
    ano_atual = agora.year

    # Consulta para contar as quantidades por marca no mês atual
    resultados = db.session.query(
        Transacao.marca_fralda_entregue,
        func.sum(Transacao.quantidade).label('total_quantidade')
    ).filter(
        func.month(Transacao.data_retirada) == mes_atual,
        func.year(Transacao.data_retirada) == ano_atual
    ).group_by(
        Transacao.marca_fralda_entregue
    ).all()

    # Calcular a média ponderada das marcas de fraldas
    total_fraldas = sum(result.total_quantidade for result in resultados)
    if total_fraldas == 0:
        return None  # Evitar divisão por zero

    # Criar um dicionário para armazenar a média sugerida por marca
    media_sugerida = defaultdict(float)
    for result in resultados:
        marca = result.marca_fralda_entregue
        quantidade = result.total_quantidade
        media_sugerida[marca] = (quantidade / total_fraldas) * 100

    return media_sugerida
def sugerir_media_mensal_tamanho():
    agora = datetime.now()
    mes_atual = agora.month
    ano_atual = agora.year

    # Consulta para contar as quantidades por tamanho de fralda no mês atual
    resultados = db.session.query(
        Transacao.tamanho_fralda,
        func.sum(Transacao.quantidade).label('total_quantidade')
    ).filter(
        func.month(Transacao.data_retirada) == mes_atual,
        func.year(Transacao.data_retirada) == ano_atual
    ).group_by(
        Transacao.tamanho_fralda
    ).all()

    # Calcular a média ponderada dos tamanhos de fraldas
    total_fraldas = sum(result.total_quantidade for result in resultados)
    if total_fraldas == 0:
        return None  # Evitar divisão por zero

    # Criar um dicionário para armazenar a média sugerida por tamanho de fralda
    media_sugerida = defaultdict(float)
    for result in resultados:
        tamanho = result.tamanho_fralda
        quantidade = result.total_quantidade
        media_sugerida[tamanho] = (quantidade / total_fraldas) * 100

    return media_sugerida
# Rota para receber os dados e processá-los temporariamente
from datetime import date
@app.route('/geraDocumento.html', methods=['GET', 'POST'])
@login_required
def mostra_gera_documento():
    impedir_repeticao = session.get('impedir_repeticao')
    nomeAutorizador = ""
    nomeFantasia = ""
    quantidade_pego = 0
    tipo_usuario = session.get('tipo_usuario')
    data_atual = datetime.now()
    if 'nomeAutorizador' in session:
        nomeAutorizador = session['nomeAutorizador']
    elif 'nomeFantasia' in session: 
        nomeFantasia = session['nomeFantasia']
        
    if request.method == 'POST':
        cpfPesquisado = Beneficiado.cpfPesquisado
        quantidade_pego = int(request.form.get('quantidade'))
        beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
        data_final= beneficiado.data_final
        if current_user.tipo_usuario == '1':  # Verifica se é uma farmácia
            farmacia = Farmacia.query.filter_by(id_usuario=current_user.id).first()
            if not farmacia:
                flash('Farmácia não encontrada.', 'error')
                return redirect(url_for('login'))
        else:
            flash('Acesso não autorizado.', 'error')
            return redirect(url_for('login'))
        if data_atual<= data_final:
            print("entra na data atual")
            if impedir_repeticao == 1:
                beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
                if beneficiado and farmacia:
                    quantidade_liberada = beneficiado.quantidade_liberada
                    quantidade_restante = beneficiado.quantidade_restante
                    print(f"Current quantidadeTotal: {farmacia.quantidade}")
                    print(f"quantidade_pego: {quantidade_pego}")
                    
                    # Verifica se a última atualização foi há mais de 30 dias
                    ultima_atualizacao = session.get('ultima_atualizacao')
                    agora = datetime.now()
                    if ultima_atualizacao:
                        ultima_atualizacao = datetime.strptime(ultima_atualizacao, '%Y-%m-%d %H:%M:%S.%f')
                        if (agora - ultima_atualizacao).days >= 30:
                            farmacia.quantidade = 0
                            flash('Quantidade total zerada após 30 dias.', 'info')
                    
                    quantidade = farmacia.quantidade + quantidade_pego
                    print(quantidade)
                    farmacia.quantidade = quantidade
                    print(f"New quantidadeTotal: {quantidade}")
                    marca = beneficiado.marca_fralda
                    if quantidade_restante == quantidade_liberada:
                        quantidade_restante = quantidade_liberada - quantidade_pego
                    if quantidade_restante == 0:
                        flash('Quantidade restante é igual a zero', 'error')
                        return render_template('visuF.html', farmacia=farmacia, beneficiado=beneficiado, quantidade_pego=quantidade_pego, quantidade_restante=quantidade_restante, nomeFantasia=nomeFantasia, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)
                    if quantidade_pego > quantidade_restante:
                        flash('Está pegando mais fraldas do que tem', 'error')
                        return render_template('visuF.html', farmacia=farmacia, beneficiado=beneficiado, quantidade_pego=quantidade_pego, quantidade_restante=quantidade_restante, nomeFantasia=nomeFantasia, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)
                    if quantidade_pego <= quantidade_restante:
                        quantidade_restante -= quantidade_pego
                    
                    tamanho = beneficiado.tamanho_liberado
                    beneficiado.quantidade_restante = quantidade_restante
                    
                    media_mensal = sugerir_media_mensal_marca()
                    media_marca = media_mensal.get(marca, 0.0) if media_mensal else 0.0
                    media_marca = Decimal(media_mensal.get(marca, 0.0)).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP) if media_mensal else Decimal('0.0')
                    media_sugerida_tamanho = sugerir_media_mensal_tamanho()
                    media_tamanho = Decimal(media_sugerida_tamanho.get(tamanho, 0.0)).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP) if media_sugerida_tamanho else Decimal('0.0')

                    hora_atual = datetime.now()
                    
                    nova_transacao = Transacao(
                        farmacia_id=farmacia.id,
                        nome_beneficiado=beneficiado.nome_beneficiado,
                        cpf_beneficiado=beneficiado.cpf_beneficiado,
                        quantidade=quantidade_pego,
                        marca_fralda_entregue=marca,
                        tamanho_fralda=tamanho,
                        quantidade_total=farmacia.quantidade,
                        media_tamanho_fralda=media_tamanho,
                        data_retirada=hora_atual
                    )

                    db.session.add(nova_transacao)
                    db.session.commit()

                    # Atualiza a data da última atualização na sessão
                    session['ultima_atualizacao'] = str(agora)
                    
                    session['farmacia'] = farmacia.id
                    session['beneficiado'] = beneficiado.id
                    session['hora_atual'] = hora_atual
                    session['quantidade_pego'] = quantidade_pego
                    session['quantidade_restante'] = quantidade_restante
                    session['nomeFantasia'] = nomeFantasia
                    session['nomeAutorizador'] = nomeAutorizador
                    session['tipo_usuario'] = tipo_usuario

                    return redirect(url_for('mostra_dados_documento'))
        if data_atual>= data_final or beneficiado.ativo=="0":
            flash("Data vencida ou sem autorização")
            return redirect(url_for('mostra_gera_documento'))
    flash('Beneficiado não encontrado ou farmácia não identificada.', 'error')
    return redirect(url_for('portal'))

def documentos_por_beneficiado(id_beneficiado):
    beneficiado = Beneficiado.query.get(id_beneficiado)
    if not beneficiado:
        return 'Beneficiado não encontrado', 404

    documento_path = os.path.join(app.config['UPLOAD_FOLDER'], beneficiado.documento)
    licitacao_path = os.path.join(app.config['UPLOAD_FOLDER'], beneficiado.licitacao)

    return send_from_directory(os.path.dirname(documento_path), os.path.basename(documento_path)), send_from_directory(os.path.dirname(licitacao_path), os.path.basename(licitacao_path))
@app.route('/mostraDadosDocumento', methods=['GET'])
@login_required
def mostra_dados_documento():
    if 'farmacia' in session and 'beneficiado' in session:
        farmacia = Farmacia.query.get(session['farmacia'])
        beneficiado = Beneficiado.query.get(session['beneficiado'])
        hora_atual = session.get('hora_atual')
        hora_atual_str = hora_atual.strftime(" Data: %d/%m/%Y, Horario: %H:%M:%S")
        quantidade_pego = session.get('quantidade_pego')
        quantidade_restante = session.get('quantidade_restante')
        nomeFantasia = session.get('nomeFantasia')
        nomeAutorizador = session.get('nomeAutorizador')
        tipo_usuario = session.get('tipo_usuario')

        return render_template("geraDocumento.html", farmacia=farmacia, beneficiado=beneficiado, hora_atual_str=hora_atual_str, quantidade_pego=quantidade_pego, quantidade_restante=quantidade_restante, nomeFantasia=nomeFantasia, nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)
    
    flash('Dados não encontrados.', 'error')
    return redirect(url_for('mostra_gera_documento'))
@app.route('/relatorioFarmacia.html')
@login_required
def relatorioFarmacia():
    beneficiados = Beneficiado.query.all()
    
    # Inicializa as variáveis para calcular as quantidades
    total_retiradas = 0
    total_disponiveis = 0
    
    # Calcula o total de fraldas retiradas e disponíveis
    for beneficiado in beneficiados:
        total_retiradas += beneficiado.quantidade_pego if beneficiado.quantidade_pego is not None else 0
    
    quantidade_restante -= total_retiradas
    
    total_pessoas = sum(beneficiado.quantidade_pego if beneficiado.quantidade_pego is not None else 0 for beneficiado in beneficiados)

    # Calcula a quantidade total de documentos anexados
    total_documentos = sum(1 for b in beneficiados if b.documento is not None or b.licitacao is not None)

    # Passa os resultados para o template HTML
    return render_template("relatorioFarmacia.html", beneficiado=beneficiado, total_pessoas=total_pessoas, 
                           total_retiradas=total_retiradas,
                           quantidade_restante=quantidade_restante, total_documentos=total_documentos)
@app.route('/anexos')
@login_required
def documentos_por_beneficiado(id_beneficiado):
    beneficiado = Beneficiado.query.get(id_beneficiado)
    if not beneficiado:
        return 'Beneficiado não encontrado', 404

    documento_path = os.path.join(app.config['UPLOAD_FOLDER'], beneficiado.documento)
    licitacao_path = os.path.join(app.config['UPLOAD_FOLDER'], beneficiado.licitacao)

    documento_response = send_from_directory(os.path.dirname(documento_path), os.path.basename(documento_path))
    licitacao_response = send_from_directory(os.path.dirname(licitacao_path), os.path.basename(licitacao_path))

    return documento_response, licitacao_response



@app.route('/usuarioPendente.html')
@login_required
def usuarioPendente():
    if 'pendente' in session and session['pendente']:
        return render_template("uPendenteF.html")
    else:
        return "Nenhuma pendência encontrada."

@app.route('/visualizaBeneficiado.html')
@login_required
def visualizaBenef():
    cpfPesquisado=Beneficiado.cpfPesquisado # Obter o CPF pesquisado do formulário
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
    return render_template("busca.html",beneficiado=beneficiado)



#criando e configutando o sistema de login
@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

@login_manager.unauthorized_handler
def unauthorized():
    # Redireciona para a página de login quando um usuário não autorizado tenta acessar uma rota protegida
    return redirect(url_for('login'))


from flask import request, session, flash, redirect, url_for, render_template, make_response
from flask_login import login_user, login_required
from werkzeug.security import check_password_hash
from datetime import timedelta

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['login']
        senha = request.form['senha']
        remember = 'remember' in request.form
        usuario = Usuario.query.filter_by(email=email).first()
        
        if usuario and usuario.ativo :
            print("entra no primeiro if")
            if check_password_hash(usuario.senha, senha) and usuario.ativo:
                if not usuario.email_confirmed:
                    flash('Por favor, confirme seu email antes de fazer login.', 'error')
                    return redirect(url_for('login'))
                
                # Realiza o login do usuário
                login_user(usuario, remember=remember)
                flash('Login bem-sucedido!', 'success')

                # Registrar o log de login do usuário
                registrar_log_usuario(usuario, usuario.id)

                # Verifica se é o primeiro login
                if usuario.first_login:
                    return redirect(url_for('trocar_senha'))


                # Armazena informações na sessão
                session['user_id'] = usuario.id  # Armazena o ID do usuário na sessão
                session['tipo_usuario'] = usuario.tipo_usuario

                if usuario.tipo_usuario == '1':  # Usuário tipo farmácia
                    farmacia = Farmacia.query.filter_by(id_usuario=usuario.id).first()
                    if farmacia:
                        session['nomeFantasia'] = farmacia.nomeFantasia.split()[0]
                        session['cnpj'] = farmacia.cnpj  # Armazena o CNPJ da farmácia na sessão
                        session['farmacia_id'] = farmacia.id  # Armazena o ID da farmácia na sessão
                    else:
                        flash('Farmácia não encontrada para o usuário logado.', 'error')
                        return redirect(url_for('login'))
                elif usuario.tipo_usuario == '2':  # Usuário tipo autorizador
                    autorizador = Autorizador.query.filter_by(id_usuario=usuario.id).first()
                    if autorizador:
                        session['nomeAutorizador'] = autorizador.nomeAutorizador.split()[0]
                        session['cpf'] = autorizador.cpf  # Armazena o CPF do autorizador na sessão
                        session['autorizador_id'] = autorizador.id
                    else:
                        flash('Autorizador não encontrado para o usuário logado.', 'error')
                        return redirect(url_for('login'))
                elif usuario.tipo_usuario == '3':  # Usuário tipo prefeituraS
                    prefeitura = Prefeitura.query.filter_by(id_usuario=usuario.id).first()
                    if prefeitura:
                        session['nomePrefeitura'] = prefeitura.nomePrefeitura.split()[0]
                        session['cpf'] = prefeitura.cpf_prefeitura  # Armazena o CPF da prefeitura na sessão
                        session['prefeitura_id'] = prefeitura.id
                    else:
                        flash('Prefeitura não encontrada para o usuário logado.', 'error')
                        return redirect(url_for('login'))
                
                response = make_response(redirect(url_for('portal')))
                if remember:
                    session.permanent = True
                    app.permanent_session_lifetime = timedelta(days=30)
                    response.set_cookie('username', email, max_age=30*24*60*60, samesite='Lax', path='/')
                else:
                    session.permanent = False

                return response
            else:
                flash('Credenciais inválidas. Por favor, tente novamente.', 'error')
                return redirect(url_for('login'))
        else:
            flash('Usuário desativado ou sem login', 'error')
            return redirect(url_for('login'))
    else:
        username = request.cookies.get('username')
        return render_template('Login.html', username=username if username else '')


from datetime import datetime
import pytz
from sqlalchemy import func

def get_current_week():
    tz = pytz.timezone('America/Sao_Paulo')  # Substitua pelo fuso horário desejado
    return datetime.now(tz).isocalendar()[1]

def get_current_year():
    tz = pytz.timezone('America/Sao_Paulo')  # Substitua pelo fuso horário desejado
    return datetime.now(tz).year

def get_total_fraldas_mes_atual(farmacia_id):
    current_month = datetime.now().month
    current_year = datetime.now().year
    total = db.session.query(func.sum(Transacao.quantidade))\
                      .filter_by(farmacia_id=farmacia_id)\
                      .filter(func.extract('month', Transacao.data_retirada) == current_month)\
                      .filter(func.extract('year', Transacao.data_retirada) == current_year)\
                      .scalar()
    print(f"Total de fraldas entregues no mês {current_month} do ano {current_year}: {total}")
    return total or 0

def get_total_fraldas_entregues(farmacia_id):
    current_week = get_current_week()
    current_year = get_current_year()
    total = db.session.query(func.sum(Transacao.quantidade))\
                      .filter_by(farmacia_id=farmacia_id)\
                      .filter(func.week(Transacao.data_retirada, True) == current_week)\
                      .filter(func.year(Transacao.data_retirada) == current_year)\
                      .scalar()
    print(f"Total de fraldas entregues na semana {current_week} do ano {current_year}: {total}")
    return total or 0

def get_marca_mais_vendida(farmacia_id):
    current_week = get_current_week()
    current_year = get_current_year()
    marca = db.session.query(Transacao.marca_fralda_entregue, func.count(Transacao.marca_fralda_entregue).label('count'))\
                      .filter_by(farmacia_id=farmacia_id)\
                      .filter(func.week(Transacao.data_retirada, True) == current_week)\
                      .filter(func.year(Transacao.data_retirada) == current_year)\
                      .group_by(Transacao.marca_fralda_entregue)\
                      .order_by(desc('count'))\
                      .first()
    print(f"Marca mais vendida na semana {current_week} do ano {current_year}: {marca[0] if marca else 'N/A'}")
    return marca[0] if marca else "N/A"

def get_media_fraldas_semana(farmacia_id):
    current_year = get_current_year()
    semanas = db.session.query(func.week(Transacao.data_retirada, True), func.sum(Transacao.quantidade).label('total'))\
                        .filter_by(farmacia_id=farmacia_id)\
                        .filter(func.year(Transacao.data_retirada) == current_year)\
                        .group_by(func.week(Transacao.data_retirada, True))\
                        .all()
    media_semanal = [semana.total for semana in semanas]
    print(f"Média de fraldas por semana no ano {current_year}: {media_semanal}")
    return media_semanal

def get_fraldas_por_dia_semana(farmacia_id):
    current_week = get_current_week()
    current_year = get_current_year()
    dias = db.session.query(func.dayofweek(Transacao.data_retirada), func.sum(Transacao.quantidade).label('total'))\
                     .filter_by(farmacia_id=farmacia_id)\
                     .filter(func.week(Transacao.data_retirada, True) == current_week)\
                     .filter(func.year(Transacao.data_retirada) == current_year)\
                     .group_by(func.dayofweek(Transacao.data_retirada))\
                     .all()

    fraldas_por_dia = [0] * 7
    mapping = {
        1: 6,  # Domingo
        2: 0,  # Segunda-feira
        3: 1,  # Terça-feira
        4: 2,  # Quarta-feira
        5: 3,  # Quinta-feira
        6: 4,  # Sexta-feira
        7: 5   # Sábado
    }

    for dia in dias:
        fraldas_por_dia[mapping[dia[0]]] = dia[1]
    
    print(f"Fraldas por dia na semana {current_week} do ano {current_year}: {fraldas_por_dia}")
    return fraldas_por_dia

def get_media_fraldas_mes(farmacia_id):
    current_year = get_current_year()
    media_mensal = db.session.query(func.avg(Transacao.quantidade).label('avg_quantidade'))\
                             .filter_by(farmacia_id=farmacia_id)\
                             .filter(func.year(Transacao.data_retirada) == current_year)\
                             .group_by(func.month(Transacao.data_retirada))\
                             .all()
    media_mensal = round(sum(media.avg_quantidade for media in media_mensal) / len(media_mensal), 1) if media_mensal else 0
    print(f"Média de fraldas por mês no ano {current_year}: {media_mensal}")
    return media_mensal

from sqlalchemy import func, desc   
from collections import OrderedDict

# Atualize a função da rota para incluir esses dados
@app.route('/portal')
@login_required
def portal():
    tipo_usuario = session.get('tipo_usuario')
    cpf = session.get('cpf')
    cnpj = session.get('cnpj')

    if tipo_usuario == '1':
        farmacia = Farmacia.query.filter_by(cnpj=cnpj).first()
        if farmacia:
            total_fraldas_entregues = get_total_fraldas_entregues(farmacia.id)
            marca_mais_vendida = get_marca_mais_vendida(farmacia.id)
            media_fraldas_semana = get_media_fraldas_semana(farmacia.id)
            fraldas_por_dia_semana = get_fraldas_por_dia_semana(farmacia.id)
            total_fraldas_mes_atual = get_total_fraldas_mes_atual(farmacia.id)
            media_fraldas_mes = get_media_fraldas_mes(farmacia.id)  # Nova função para a média mensal

            # Convertendo media_fraldas_semana para uma lista simples de números
            media_fraldas_semana = [float(value) for value in media_fraldas_semana]

            return render_template('portal.html', tipo_usuario=tipo_usuario, 
                                   nomeFantasia=farmacia.nomeFantasia, 
                                   total_fraldas_entregues=total_fraldas_entregues, 
                                   marca_mais_vendida=marca_mais_vendida, 
                                   media_fraldas_semana=media_fraldas_semana,
                                   fraldas_por_dia_semana=fraldas_por_dia_semana,
                                   total_fraldas_mes_atual=total_fraldas_mes_atual,
                                   media_fraldas_mes=media_fraldas_mes)  # Passando a média mensal
    elif tipo_usuario == '2':
        autorizador = Autorizador.query.filter_by(cpf=cpf).first()
        if autorizador:
            # Calcular o total de beneficiados cadastrados por mês pelo autorizador logado
            beneficiados_por_mes = db.session.query(
                db.func.date_format(Beneficiado.data_cadastro, '%Y-%m').label('mes'),
                db.func.count(Beneficiado.id).label('total_beneficiados')
            ).filter(Beneficiado.id_autorizador == autorizador.id).group_by('mes').order_by('mes').all()

            total_beneficiados_mensais = {mes: total for mes, total in beneficiados_por_mes}
            total_beneficiados_geral = sum(total_beneficiados_mensais.values())

            marcas_por_mes = db.session.query(
                db.func.date_format(Beneficiado.data_cadastro, '%Y-%m').label('mes'),
                Beneficiado.marca_fralda,
                db.func.count(Beneficiado.marca_fralda).label('total')
            ).filter(Beneficiado.id_autorizador == autorizador.id).group_by('mes', Beneficiado.marca_fralda).all()

            marcas_mais_liberadas_mensais = {}
            for mes, marca, total in marcas_por_mes:
                if mes not in marcas_mais_liberadas_mensais:
                    marcas_mais_liberadas_mensais[mes] = {}
                marcas_mais_liberadas_mensais[mes][marca] = total

            marca_mais_liberada_mensal = {}
            for mes, marcas in marcas_mais_liberadas_mensais.items():
                if marcas:  # Verificar se a lista não está vazia
                    marca_mais_liberada = max(marcas, key=marcas.get)
                    marca_mais_liberada_mensal[mes] = marca_mais_liberada

            # Encontrar a marca mais liberada no total
            if marca_mais_liberada_mensal:  # Verificar se a lista não está vazia
                marca_mais_liberada_footer = max(marca_mais_liberada_mensal.values(), key=list(marca_mais_liberada_mensal.values()).count)
            else:
                marca_mais_liberada_footer = None

            # Mapeamento de códigos de meses para nomes em português
            meses_nome = {
                '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
                '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
                '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
            }

            # Ordenar os dados dos meses por ordem cronológica
            meses_ordenados = sorted(total_beneficiados_mensais.items())
            total_beneficiados_mensais_ordenado = OrderedDict(meses_ordenados)

            tamanhos_por_mes = db.session.query(
                db.func.date_format(Beneficiado.data_cadastro, '%Y-%m').label('mes'),
                Beneficiado.tamanho_liberado,
                db.func.count(Beneficiado.tamanho_liberado).label('total')
            ).filter(Beneficiado.id_autorizador == autorizador.id).group_by('mes', Beneficiado.tamanho_liberado).all()

            tamanhos_mais_liberados_mensais = {}
            for mes, tamanho, total in tamanhos_por_mes:
                if mes not in tamanhos_mais_liberados_mensais:
                    tamanhos_mais_liberados_mensais[mes] = {}
                tamanhos_mais_liberados_mensais[mes][tamanho] = total

            tamanho_mais_liberado_mensal = {}
            for mes, tamanhos in tamanhos_mais_liberados_mensais.items():
                if tamanhos:  # Verificar se a lista não está vazia
                    tamanho_mais_liberado = max(tamanhos, key=tamanhos.get)
                    tamanho_mais_liberado_mensal[mes] = tamanho_mais_liberado

            # Encontrar o tamanho mais liberado no total
            if tamanho_mais_liberado_mensal:  # Verificar se a lista não está vazia
                tamanho_mais_liberado_footer = max(tamanho_mais_liberado_mensal.values(), key=list(tamanho_mais_liberado_mensal.values()).count)
            else:
                tamanho_mais_liberado_footer = None

            total_beneficiados_mensais_formatado = {meses_nome[mes[-2:]]: total for mes, total in total_beneficiados_mensais_ordenado.items()}
            marca_mais_liberada_mensal_formatado = {meses_nome[mes[-2:]]: marca for mes, marca in marca_mais_liberada_mensal.items()}
            tamanho_mais_liberado_mensal_formatado = {meses_nome[mes[-2:]]: tamanho for mes, tamanho in tamanho_mais_liberado_mensal.items()}

            return render_template('portal.html', tipo_usuario=tipo_usuario,
                                   nomeAutorizador=autorizador.nomeAutorizador,
                                   total_beneficiados_mensais=total_beneficiados_mensais_formatado,
                                   total_beneficiados_geral=total_beneficiados_geral,
                                   marca_mais_liberada_mensal=marca_mais_liberada_mensal_formatado,
                                   marca_mais_liberada_footer=marca_mais_liberada_footer,
                                   tamanho_mais_liberado_mensal=tamanho_mais_liberado_mensal_formatado,
                                   tamanho_mais_liberado_footer=tamanho_mais_liberado_footer)
        
    elif tipo_usuario == '3':
        prefeitura = Prefeitura.query.filter_by(cpf_prefeitura=cpf).first()
        if prefeitura:
            total_farmacias_geral = Farmacia.query.count()

            # Calcular o total de fraldas entregues por mês e o total geral
            transacoes_por_mes = db.session.query(
                db.func.date_format(Transacao.data_retirada, '%Y-%m').label('mes'),
                db.func.sum(Transacao.quantidade).label('total_fraldas')
            ).group_by('mes').order_by('mes').all()

            total_fraldas_mensais = {mes: total for mes, total in transacoes_por_mes}
            total_fraldas_geral = sum(total_fraldas_mensais.values())

            # Calcular o total de beneficiados cadastrados por mês e o total geral
            beneficiados_por_mes = db.session.query(
                db.func.date_format(Beneficiado.data_cadastro, '%Y-%m').label('mes'),
                db.func.count(Beneficiado.id).label('total_beneficiados')
            ).group_by('mes').order_by('mes').all()

            total_beneficiados_mensais = {mes: total for mes, total in beneficiados_por_mes}
            total_beneficiados_geral = sum(total_beneficiados_mensais.values())

            # Calcular o total de farmácias cadastradas por mês e o total geral
            farmacias_por_mes = db.session.query(
                db.func.date_format(Farmacia.data_cadastro, '%Y-%m').label('mes'),
                db.func.count(Farmacia.id).label('total_farmacias')
            ).group_by('mes').order_by('mes').all()

            total_farmacias_mensais = {mes: total for mes, total in farmacias_por_mes}
            total_farmacias_geral = sum(total_farmacias_mensais.values())

            # Ordenar os dados dos meses por ordem cronológica
            meses_ordenados_fraldas = sorted(total_fraldas_mensais.items())
            total_fraldas_mensais_ordenado = OrderedDict(meses_ordenados_fraldas)

            meses_ordenados_beneficiados = sorted(total_beneficiados_mensais.items())
            total_beneficiados_mensais_ordenado = OrderedDict(meses_ordenados_beneficiados)

            meses_ordenados_farmacias = sorted(total_farmacias_mensais.items())
            total_farmacias_mensais_ordenado = OrderedDict(meses_ordenados_farmacias)

            # Converter os meses para nomes completos em português
            meses_nome = {
                '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março', '04': 'Abril',
                '05': 'Maio', '06': 'Junho', '07': 'Julho', '08': 'Agosto',
                '09': 'Setembro', '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
            }

            total_fraldas_mensais_formatado = {meses_nome[mes[-2:]]: total for mes, total in total_fraldas_mensais_ordenado.items()}
            total_beneficiados_mensais_formatado = {meses_nome[mes[-2:]]: total for mes, total in total_beneficiados_mensais_ordenado.items()}
            total_farmacias_mensais_formatado = {meses_nome[mes[-2:]]: total for mes, total in total_farmacias_mensais_ordenado.items()}

            return render_template('portal.html', tipo_usuario=tipo_usuario, 
                                   nomePrefeitura=prefeitura.nomePrefeitura,
                                   total_farmacias_mensais=total_farmacias_mensais_formatado,
                                   total_farmacias_geral=total_farmacias_geral,
                                   total_fraldas_mensais=total_fraldas_mensais_formatado,
                                   total_fraldas_geral=total_fraldas_geral,
                                   total_beneficiados_mensais=total_beneficiados_mensais_formatado,
                                   total_beneficiados_geral=total_beneficiados_geral)
    
    flash('Usuário não encontrado', 'danger')
    return redirect(url_for('login'))


@app.route('/perfil')
@login_required
def perfil():
    tipo_usuario = session.get('tipo_usuario')
    cpf = session.get('cpf')
    cnpj = session.get('cnpj')

    if tipo_usuario == '1':
        farmacia = Farmacia.query.filter_by(cnpj=cnpj).first()
        if farmacia:
            return render_template('perfil.html', tipo_usuario=tipo_usuario, 
                                   nomeFantasia=farmacia.nomeFantasia, 
                                   cnpj=farmacia.cnpj, 
                                   localizacao=f"{farmacia.logradouro}, {farmacia.numero}, {farmacia.bairro}, {farmacia.cidade} - {farmacia.estado}", 
                                   email=farmacia.usuario.email,
                                   farmacia_id=farmacia.id)  # Passando o ID da farmácia
    elif tipo_usuario == '2':
        autorizador = Autorizador.query.filter_by(cpf=cpf).first()
        if autorizador:
            return render_template('perfil.html', tipo_usuario=tipo_usuario, 
                                   nomeAutorizador=autorizador.nomeAutorizador, 
                                   cpf=autorizador.cpf, 
                                   email=autorizador.usuario.email,
                                   autorizador_id=autorizador.id)  # Passando o ID do autorizador
    elif tipo_usuario == '3':
        prefeitura = Prefeitura.query.filter_by(cpf_prefeitura=cpf).first()
        if prefeitura:
            return render_template('perfil.html', tipo_usuario=tipo_usuario, 
                                   nomePrefeitura=prefeitura.nomePrefeitura, 
                                   cpf=prefeitura.cpf_prefeitura, 
                                   email=prefeitura.usuario.email,
                                   prefeitura_id=prefeitura.id)  # Passando o ID da prefeitura
    else:
        flash('Tipo de usuário não reconhecido', 'danger')
        return redirect(url_for('farmacia.pagina_farmacia'))

    flash('Usuário não encontrado', 'danger')
    return redirect(url_for('farmacia.pagina_farmacia'))



@app.route('/trocar_email', methods=['GET', 'POST'])
@login_required
def trocar_email():
    nomeAutorizador = ""
    nomeFantasia = ""
    nomePrefeitura =""
    tipo_usuario = session.get('tipo_usuario')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    elif 'nomeAutorizador' in session:
        nomeAutorizador = session['nomeAutorizador']
    elif 'nomeFantasia' in session: 
        nomeFantasia = session ['nomeFantasia']
    
    if request.method == 'POST':
        current_email = request.form['current_email']
        new_email = request.form['new_email']

        if current_user.email != current_email:
            flash('O email atual não corresponde ao email do usuário logado.', 'danger')
            return redirect(url_for('trocar_email'))

        token = s.dumps(new_email, salt='email-confirm')

        msg = Message('Confirme seu novo endereço de email',
                      sender='seu_email@gmail.com',
                      recipients=[new_email])
        link = url_for('confirmar_email', token=token, _external=True)
        msg.body = f'Por favor, clique no link para confirmar seu novo endereço de email: {link}'
        mail.send(msg)

        flash('Um email de confirmação foi enviado para o novo endereço.', 'info')
        return redirect(url_for('trocar_email'))
    return render_template('trocar_dados.html',page='trocar_email', tipo_usuario=tipo_usuario, nomePrefeitura=nomePrefeitura, nomeAutorizador=nomeAutorizador, nomeFantasia=nomeFantasia)

@app.route('/confirmar_email/<token>')
def confirmar_email(token):
    try:
        new_email = s.loads(token, salt='email-confirm', max_age=3600)
    except:
        flash('O link de confirmação é inválido ou expirou.', 'danger')
        return redirect(url_for('login'))

    user = Usuario.query.get(current_user.id)
    user.email = new_email
    db.session.commit()

    flash('Seu endereço de email foi atualizado com sucesso!', 'success')
    return redirect(url_for('login'))  # Redireciona para a página de perfil ou onde você desejar

@app.route('/trocar_senha', methods=['GET', 'POST'])
@login_required

def trocar_senha():

    nomeAutorizador = ""
    nomeFantasia = ""
    nomePrefeitura =""
    tipo_usuario = session.get('tipo_usuario')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    elif 'nomeAutorizador' in session:
        nomeAutorizador = session['nomeAutorizador']
    elif 'nomeFantasia' in session: 
        nomeFantasia = session ['nomeFantasia']

    if request.method == 'POST':
        nova_senha = request.form['nova_senha']
        confirmar_senha = request.form['confirmar_senha']

        if nova_senha != confirmar_senha:
            flash('As senhas não coincidem. Por favor, tente novamente.', 'error')
        elif not re.match(r'^(?=.*[a-zA-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$', nova_senha):
            flash('A senha deve ter pelo menos 8 caracteres, incluindo letras, números e um caractere especial.', 'error')
        else:
            usuario = current_user
            usuario.senha = generate_password_hash(nova_senha)
            usuario.first_login = False  
            db.session.commit()         

            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('login'))


    return render_template('trocar_dados.html', page='trocar_senha', first_login=current_user.first_login, tipo_usuario=tipo_usuario, nomeAutorizador=nomeAutorizador, nomeFantasia=nomeFantasia, nomePrefeitura=nomePrefeitura)

@app.route('/visu')
@login_required
def visu():
    return render_template("visu.html")





@app.route('/visualizaBeneficiadoUm')
@login_required
def visualizaBeneficiadoUm():
    cpfPesquisado=Beneficiado.cpfPesquisado # Obter o CPF pesquisado do formulário
    beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=cpfPesquisado).first()
    return render_template("visu.html", page = 'visuA', beneficiado=beneficiado)

@app.route("/VisuBeneficiadoPref.html")
@login_required
def VisuBeneficiadoPref():
    return render_template("visu.html", page = 'visuP')

@login_manager.unauthorized_handler
def unauthorized():
    # Redireciona para a página de login quando um usuário não autorizado tenta acessar uma rota protegida
    return redirect(url_for('login'))
@app.route('/logout', methods=['GET', 'POST'])
def logout():
    if current_user.is_authenticated:
        username = request.cookies.get('username')  # Preserve o nome de usuário
      
        logout_user()
        session.clear()  # Limpa todos os dados da sessão
        flash('Você foi desconectado com sucesso.', 'success')
        response = make_response(redirect(url_for('login')))
        if username:
            response.set_cookie('username', username, max_age=30*24*60*60, samesite='Lax', path='/')
            print(f'Resetting cookie: username={username}')  # Log para verificar
        return response
    else:
        flash('Você não está logado.', 'error')
        return redirect(url_for('login'))

def registro_de_pesquisa_cnpj(cnpj_pesquisado):
    data_pesquisa = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    status = 'encontrado' if cnpj_pesquisado else 'não encontrado'
    descricao_log = f'Pesquisa realizada:\n- Data: {data_pesquisa}\n- CPF Pesquisado: {cnpj_pesquisado}\n- Status: {status}\n'
    
    try:
        with open('log.txt', 'a') as arquivo_log:
            arquivo_log.write(descricao_log)
    except Exception as e:
        print(f"Erro ao escrever no arquivo de log: {e}")

#procurar a farmacia pelo cnpj
@app.route('/pesquisarPfarmacia', methods=['GET', 'POST'])
@login_required
def pesquisarPfarmacia():
    tipo_usuario = session.get('tipo_usuario')
    nomePrefeitura = session.get('nomePrefeitura')

    if request.method == 'POST':
        cnpj_pesquisado = request.form.get('cnpj_pesquisado')
        session['cnpj_pesquisado'] = cnpj_pesquisado
        if not cnpj_pesquisado:
            flash('CNPJ não fornecido.', 'error')
            return render_template('buscaPrefeitura.html', page='buscaPFarmacia', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

        try:
            farmacia = Farmacia.query.filter_by(cnpj=cnpj_pesquisado).first()

            if not farmacia:
                flash('Farmácia não encontrada.', 'error')
                return render_template('buscaPrefeitura.html', page='buscaPFarmacia', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            registro_de_pesquisa_cnpj(cnpj_pesquisado)

            transacoes = Transacao.query.filter_by(farmacia_id=farmacia.id).all()

            if not transacoes:
                flash('Nenhuma transação encontrada para esta farmácia.', 'info')
                return render_template(
                    "visualizarPrefeitura.html", page = 'visuPfarmacia',
                    marca_fralda_entregue=None,
                    media_marca_fralda={},
                    transacoes=[],
                    farmacia=farmacia,
                    cnpj_pesquisado=cnpj_pesquisado,
                    nomePrefeitura=nomePrefeitura,
                    tipo_usuario=tipo_usuario
                )

            total_quantidade = sum(transacao.quantidade for transacao in transacoes)

            if total_quantidade > 0:
                quantidade_por_marca = defaultdict(int)
                for transacao in transacoes:
                    quantidade_por_marca[transacao.marca_fralda_entregue] += transacao.quantidade

                media_marca_fralda = {marca: (quantidade / total_quantidade) * 100 for marca, quantidade in quantidade_por_marca.items()}
            else:
                media_marca_fralda = {}

            marca_fralda_entregue = max(quantidade_por_marca, key=quantidade_por_marca.get, default=None)

            return render_template(
                "visualizarPrefeitura.html", page ='visuPfarmacia',
                marca_fralda_entregue=marca_fralda_entregue,
                media_marca_fralda=media_marca_fralda,
                transacoes=transacoes,
                farmacia=farmacia,
                cnpj_pesquisado=cnpj_pesquisado,
                nomePrefeitura=nomePrefeitura,
                tipo_usuario=tipo_usuario
            )

        except Exception as e:
            flash(f'Ocorreu um erro ao consultar o banco de dados: {str(e)}', 'error')
            return render_template('buscaPrefeitura.html', page='BuscaPFarmacia', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

    else:
        flash('Método HTTP inválido para esta rota.', 'error')
        return render_template('buscaPrefeitura.html', page='BuscaPFarmacia', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)


@app.route('/pesquisarPrefeitura', methods=['GET', 'POST'])
@login_required
def pesquisarPrefeitura():
    tipo_usuario = session.get('tipo_usuario')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    
    try:
        if request.method == 'POST':
            # Obter o CNPJ pesquisado do formulário
            cpf_pesquisado = request.form.get('cpf_pesquisado')
            session['cpf_pesquisado'] = cpf_pesquisado
            if not cpf_pesquisado:
                flash('CNPJ não fornecido.', 'error')
                return render_template('buscaPrefeitura.html')

            try:
                # Consultar a farmácia pelo CNPJ pesquisado
                prefeitura = Prefeitura.query.filter_by(cpf=cpf_pesquisado).first()

                if not prefeitura:
                    flash('Farmácia não encontrada.', 'error')
                    return render_template('buscaPrefeitura.html')
                return render_template(
                    "visuPrefeitura", page = 'visuPfarmacia',
                    cpf_pesquisado=cpf_pesquisado, nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario
                )


            except Exception as e:
                flash(f'Ocorreu um erro ao consultar o banco de dados: {str(e)}', 'error')
                return render_template('buscaPfarmacia.html', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

        else:
            flash('Método HTTP inválido para esta rota.', 'error')
            return render_template('buscaPrefeitura.html', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

    except Exception as e:
        flash(f'Ocorreu um erro ao processar a pesquisa: {str(e)}', 'error')
        return render_template("prefeitura.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

# procurar o autorizador pelo cpf
def is_valid_cpf(cpf: str) -> bool:
    cpf = ''.join(filter(str.isdigit, cpf))
    if len(cpf) != 11:
        return False
    if cpf in [str(i) * 11 for i in range(10)]:
        return False
    sum1 = sum(int(cpf[i]) * (10 - i) for i in range(9))
    digit1 = (sum1 * 10 % 11) % 10
    sum2 = sum(int(cpf[i]) * (11 - i) for i in range(10))
    digit2 = (sum2 * 10 % 11) % 10
    return cpf[-2:] == f'{digit1}{digit2}'

from datetime import datetime, timedelta
from sqlalchemy import extract, func
from pytz import timezone
@app.route('/pesquisarPautorizador', methods=['POST'])
def pesquisarPautorizador():
    tipo_usuario = session.get('tipo_usuario')
    nomePrefeitura = session.get('nomePrefeitura')  # Obtém o nome da prefeitura da sessão, se estiver presente

    try:
        if request.method == 'POST':
            cpf_pesquisado = request.form['cpf_pesquisado']
            session['cpf_pesquisado'] = cpf_pesquisado

            if not is_valid_cpf(cpf_pesquisado):
                flash('Insira um CPF válido.', 'error')
                return render_template('buscaPrefeitura.html', page='BuscaPAutorizador', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

            autorizador = Autorizador.query.filter_by(cpf=cpf_pesquisado).first()
            registro_de_pesquisa(cpf_pesquisado, autorizador is not None)
            
            if autorizador:
                db.session.commit()  # Salvar a atualização no banco de dados

                # Calculando total por mês e média anual
                current_month = datetime.now().month
                current_year = datetime.now().year

                total_por_mes = TransacaoA.query.filter(
                    extract('month', TransacaoA.data_retirada) == current_month,
                    extract('year', TransacaoA.data_retirada) == current_year,
                    TransacaoA.autorizador_id == autorizador.id
                ).with_entities(func.sum(TransacaoA.quantidadeBeneficiado)).scalar() or 0

                total_por_ano = TransacaoA.query.filter(
                    extract('year', TransacaoA.data_retirada) == current_year,
                    TransacaoA.autorizador_id == autorizador.id
                ).with_entities(func.sum(TransacaoA.quantidadeBeneficiado)).scalar() or 0

                media_anual = total_por_ano / 12 if total_por_ano > 0 else 0

                # Calculando média mensal
                tz = timezone('America/Sao_Paulo')
                now = datetime.now(tz)
                first_day_of_month = now.replace(day=1)
                last_day_of_month = now.replace(day=1, month=now.month % 12 + 1) - timedelta(days=1)

                total_mes_atual = TransacaoA.query.filter(
                    TransacaoA.data_retirada.between(first_day_of_month, last_day_of_month),
                    TransacaoA.autorizador_id == autorizador.id
                ).with_entities(func.sum(TransacaoA.quantidadeBeneficiado)).scalar() or 0

                media_mensal = total_mes_atual / now.day if now.day > 0 else 0

                # Renderizar o template visuPautorizador.html e passar os dados necessários
                return render_template("visualizarPrefeitura.html", page = 'visuPautorizador', 
                                       autorizador=autorizador, 
                                       cpf_pesquisado=cpf_pesquisado, 
                                       nomePrefeitura=nomePrefeitura, 
                                       tipo_usuario=tipo_usuario, 
                                       total_por_mes=total_por_mes, 
                                       media_anual=media_anual, 
                                       media_mensal=media_mensal)

            else:
                flash('Beneficiado não encontrado.', 'error')
                return render_template('buscaPrefeitura.html', page = 'BuscaPAutorizador', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

        else:
            flash('Método HTTP inválido para esta rota.', 'error')
            return render_template('prefeitura.html', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

    except Exception as e:
        app.logger.error(f"Erro durante a pesquisa de autorizador: {str(e)}")
        flash('Ocorreu um erro. Por favor, tente novamente.', 'error')
        return render_template("prefeitura.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)
#procurar o beneficiado pelo cpf
@app.route('/pesquisarPbeneficiado', methods=['POST'])
def pesquisarPbeneficiado():
    tipo_usuario = session.get('tipo_usuario')
    nomePrefeitura = session.get('nomePrefeitura', 'Desconhecido')

    try:
        if request.method == 'POST':
            # Obtenha o CPF pesquisado do formulário
            cpf_pesquisado = request.form.get('cpf_pesquisado', '')
            Beneficiado.cpfPesquisado = cpf_pesquisado
            beneficiado = Beneficiado.query.filter_by(cpf_beneficiado=Beneficiado.cpfPesquisado).first()

            # Registre a pesquisa
            registro_de_pesquisa(Beneficiado.cpfPesquisado, beneficiado is not None)
            
            if beneficiado:
                autorizador = Autorizador.query.filter_by(id_usuario=beneficiado.id_autorizador).first()

                # Verifica se o autorizador foi encontrado
                if autorizador:
                    nome_autorizador = autorizador.nomeAutorizador
                    cpf_autorizador = autorizador.cpf
                else:
                    nome_autorizador = 'Desconhecido'
                    cpf_autorizador = 'Desconhecido'
                
                db.session.commit()  # Salvar a atualização no banco de dados

                # Renderizar o template visualizaFarmaciaUm.html e passar os dados do beneficiado como argumentos
                return render_template(
                    "visualizarPrefeitura.html",
                    page='visuPbeneficiado',
                    beneficiado=beneficiado,
                    cpfPesquisado=Beneficiado.cpfPesquisado,
                    nome_autorizador=nome_autorizador,
                    cpf_autorizador=cpf_autorizador,
                    nomePrefeitura=nomePrefeitura,
                    tipo_usuario=tipo_usuario
                )
            else:
                flash('Beneficiado não encontrado.', 'error')
                return render_template(
                    'buscaPrefeitura.html',
                    page='buscaPBeneficiado',
                    nomePrefeitura=nomePrefeitura,
                    tipo_usuario=tipo_usuario
                )
        else:
            flash('Método HTTP inválido para esta rota.', 'error')
            return render_template(
                'prefeitura.html',
                nomePrefeitura=nomePrefeitura,
                tipo_usuario=tipo_usuario
            )
    except Exception as e:
        flash(f'Ocorreu um erro: {e}', 'error')
        return render_template(
            "portal.html",
            nomePrefeitura=nomePrefeitura,
            tipo_usuario=tipo_usuario
        )

        

#gerando uma planilha 
# Rota para gerar e baixar a planilha
from sqlalchemy import extract
from flask import request
from sqlalchemy import or_
import pytz
from flask import Flask, request, make_response, render_template
from flask_login import current_user, login_required
from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import or_, extract
import pytz

@app.route('/gerar_planilhaF', methods=['GET', 'POST'])
@login_required
def gerar_planilhaF():
    if request.method == 'POST':
        ano_filter = request.form.get('ano-filter')
        mes_filter = request.form.get('mes-filter')

        app.logger.info(f"Filtros recebidos: ano_filter={ano_filter}, mes_filter={mes_filter}")

        usuario = current_user
        farmacia = usuario.farmacia

        if not farmacia:
            return "Usuário não associado a uma farmácia", 400

        query = db.session.query(Farmacia, Transacao).join(Transacao, Farmacia.id == Transacao.farmacia_id)
        query = query.filter(Farmacia.id == farmacia.id)

        if ano_filter:
            try:
                ano_filter = int(ano_filter)
                query = query.filter(extract('year', Transacao.data_retirada) == ano_filter)
            except ValueError:
                return "Ano inválido", 400

        if mes_filter:
            try:
                mes_filter = int(mes_filter)
                query = query.filter(extract('month', Transacao.data_retirada) == mes_filter)
            except ValueError:
                return "Mês inválido", 400

        app.logger.info(f"Consulta SQL gerada: {query}")

        dados = query.all()

        app.logger.info(f"Dados encontrados na consulta: {dados}")

        if not dados:
            flash("Nenhum dados para baixar")
            return redirect(url_for('farmacia.pagina_farmacia'))

        wb = Workbook()
        ws = wb.active
        ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"

        ws.append(['Nome Fantasia', 'CNPJ', 'Quantidade', 'Marca', 'Tamanho', 'Quantidade Total', 'Data'])

        for farmacia, transacao in dados:
            data_retirada = transacao.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
            ws.append([
                farmacia.nomeFantasia,
                farmacia.cnpj,
                transacao.quantidade,
                transacao.marca_fralda_entregue,
                transacao.tamanho_fralda,
                farmacia.quantidade,
                data_retirada.strftime('%Y-%m-%d %H:%M:%S')
            ])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

        return response

    return render_template('panilhaFarmaciaF.html')


@app.route('/mostrar_planilhaF', methods=['GET'])
@login_required
def mostrar_planilhaF():
    ano_filter = request.args.get('ano')
    mes_filter = request.args.get('mes')
    name_filter = request.args.get('name-filter')
    
    app.logger.info(f"Filtros: ano_filter={ano_filter}, mes_filter={mes_filter}, name_filter={name_filter}")

    usuario = current_user
    farmacia = usuario.farmacia

    if not farmacia:
        return "Usuário não associado a uma farmácia", 400
    
    query = db.session.query(Farmacia, Transacao).join(Transacao, Farmacia.id == Transacao.farmacia_id)
    query = query.filter(Farmacia.id == farmacia.id)
    
    if name_filter:
        query = query.filter(
            or_(
                Transacao.nome_beneficiado.ilike(f'%{name_filter}%'),
                Farmacia.cnpj.ilike(f'%{name_filter}%')
            )
        )
    
    if ano_filter:
        try:
            ano_filter = int(ano_filter)
            query = query.filter(extract('year', Transacao.data_retirada) == ano_filter)
        except ValueError:
            return "Ano inválido", 400

    if mes_filter:
        try:
            mes_filter = int(mes_filter)
            query = query.filter(extract('month', Transacao.data_retirada) == mes_filter)
        except ValueError:
            return "Mês inválido", 400

    app.logger.info(f"Consulta SQL gerada: {query}")

    dados = query.all()

    app.logger.info(f"Dados encontrados na consulta: {dados}")

    if not dados:
        flash("Nenhum dado encontrado para os filtros especificados.")
        return render_template('farmacia.html')

    # Cria um novo arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"

    # Cabeçalho da planilha
    # Itera sobre os dados para preencher a planilha
    for farmacia, transacao in dados:
        data_retirada = transacao.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
        ws.append([
            farmacia.nomeFantasia,
            farmacia.cnpj,
            transacao.nome_beneficiado,
            transacao.cpf_beneficiado,
            transacao.marca_fralda_entregue,
            transacao.tamanho_fralda,
            transacao.quantidade,
            transacao.quantidade_total,
            data_retirada.strftime('%d-%m-%Y %H:%M:%S')
        ])

    # Salva o arquivo Excel em memória
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Prepara a resposta para download
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

    return response



@app.route('/gerar_planilha', methods=['GET', 'POST'])
def gerar_planilha():
    if request.method == 'POST':
        # Obter os filtros de nome, ano e mês do formulário
        name_filter = request.form.get('name-filter')
        ano_filter = request.form.get('ano-filter')
        mes_filter = request.form.get('mes-filter')

        app.logger.info(f"Filtros recebidos: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

        # Iniciar a query na tabela 'Farmacia' e 'Transacao'
        query = db.session.query(Farmacia, Transacao).join(Transacao, Farmacia.id == Transacao.farmacia_id)

        # Aplicar filtro pelo nome, se fornecido
        if name_filter:
            query = query.filter(
                or_(
                    Farmacia.nomeFantasia.ilike(f'%{name_filter}%'),
                    Farmacia.cnpj.ilike(f'%{name_filter}%')
                )
            )

        # Aplicar filtro pelo ano, se fornecido
        if ano_filter:
            try:
                ano_filter = int(ano_filter)
                query = query.filter(extract('year', Transacao.data_retirada) == ano_filter)
            except ValueError:
                return "Ano inválido", 400

        # Aplicar filtro pelo mês, se fornecido
        if mes_filter:
            try:
                mes_filter = int(mes_filter)
                query = query.filter(extract('month', Transacao.data_retirada) == mes_filter)
            except ValueError:
                return "Mês inválido", 400

        app.logger.info(f"Consulta SQL gerada: {query}")

        # Executar a consulta e obter os resultados
        dados = query.all()

        app.logger.info(f"Dados encontrados na consulta: {dados}")

        # Criar uma nova planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"  # Usar 'Todos' se nenhum filtro especificado

        # Adicionar a linha de cabeçalho na planilha
        ws.append(['Nome Fantasia', 'CNPJ', 'Nome do beneficiado contemplado','CPF do beneficiado contemplado',  'Marca', 'Tamanho', 'Quantidade Pega', 'Quantidade','Data'])

        for farmacia, transacao in dados:
            data_retirada = transacao.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
            ws.append([
                    farmacia.nomeFantasia,
                    farmacia.cnpj,
                    transacao.nome_beneficiado,
                    transacao.cpf_beneficiado,          
                    transacao.marca_fralda_entregue,
                    transacao.tamanho_fralda,   
                    transacao.quantidade,
                    transacao.quantidade_total,
                    data_retirada.strftime('%d-%m-%Y %H:%M:%S')
                ])

        # Salvar a planilha em um objeto BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Criar uma resposta com o arquivo Excel
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

        return response
    
    # Renderizar o template HTML se o método for GET
    return render_template('planilhaFarmacia.html')
from sqlalchemy import extract

@app.route('/mostrar_planilha', methods=['GET'])
@login_required
def mostrar_planilha():
    # Obter os filtros de nome, ano e mês dos parâmetros da URL
    name_filter = request.args.get('name-filter')
    ano_filter = request.args.get('ano')
    mes_filter = request.args.get('mes')

    app.logger.info(f"Filtros: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

    # Construir a consulta SQL base
    query = db.session.query(Farmacia, Transacao).join(Transacao, Farmacia.id == Transacao.farmacia_id)

    # Aplicar filtro pelo nome, se fornecido
    if name_filter:
        query = query.filter(
            or_(
                Transacao.nome_beneficiado.ilike(f'%{name_filter}%'),
                Farmacia.cnpj.ilike(f'%{name_filter}%')
            )
        )

    # Aplicar filtro pelo ano, se fornecido
    if ano_filter:
        try:
            ano_filter = int(ano_filter)
            query = query.filter(extract('year', Transacao.data_retirada) == ano_filter)
        except ValueError:
            return "Ano inválido", 400

    # Aplicar filtro pelo mês, se fornecido
    if mes_filter:
        try:
            mes_filter = int(mes_filter)
            query = query.filter(extract('month', Transacao.data_retirada) == mes_filter)
        except ValueError:
            return "Mês inválido", 400

    app.logger.info(f"Consulta SQL gerada: {query}")

    # Executar a consulta e obter os resultados
    dados = query.all()

    app.logger.info(f"Dados encontrados na consulta: {dados}")

    # Criar uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"  # Usar 'Todos' se nenhum filtro especificado

    # Adicionar a linha de cabeçalho na planilha
    ws.append(['Nome Fantasia', 'CNPJ', 'Quantidade', 'Marca', 'Tamanho','Quantidade Total', 'Media de Marca de fralda', 'Media Tamanho de fralda', 'Data'])

    # Adicionar os dados na planilha
    for farmacia, transacao in dados:
        data_retirada = transacao.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
        ws.append([farmacia.nomeFantasia, farmacia.cnpj, transacao.quantidade, transacao.marca_fralda_entregue, transacao.tamanho_fralda, farmacia.quantidade, transacao.media_marca_fralda, transacao.media_tamanho_fralda, transacao.data_retirada.strftime('%Y-%m-%d %H:%M:%S')])

    # Salvar a planilha em um objeto BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Criar uma resposta com o arquivo Excel
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

    return response


# Route to generate and download the spreadsheet for Beneficiados
from dateutil.relativedelta import relativedelta

@app.route('/gerar_planilhaB', methods=['GET', 'POST'])
@login_required
def gerar_planilhaB():
    if request.method == 'POST':
        # Obter os filtros de nome, ano e mês do formulário
        name_filter = request.form.get('name-filter')
        ano_filter = request.form.get('ano-filter')
        mes_filter = request.form.get('mes-filter')

        app.logger.info(f"Filtros recebidos: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

        # Iniciar a query na tabela 'Beneficiado'
        query = db.session.query(Beneficiado)

        # Aplicar filtro pelo nome ou CPF, se fornecido
        if name_filter:
            query = query.filter(
                or_(
                    Beneficiado.nome_beneficiado.ilike(f'%{name_filter}%'),
                    Beneficiado.cpf_beneficiado.ilike(f'%{name_filter}%')
                )
            )

        # Aplicar filtro pelo ano, se fornecido
        if ano_filter:
            try:
                ano_filter = int(ano_filter)
                query = query.filter(extract('year', Beneficiado.data_inicio) == ano_filter)
            except ValueError:
                return "Ano inválido", 400

        # Aplicar filtro pelo mês, se fornecido
        if mes_filter:
            try:
                mes_filter = int(mes_filter)
                query = query.filter(extract('month', Beneficiado.data_inicio) == mes_filter)
            except ValueError:
                return "Mês inválido", 400

        app.logger.info(f"Consulta SQL gerada: {query}")

        # Executar a consulta e obter os resultados
        beneficiados = query.all()

        app.logger.info(f"Dados encontrados na consulta: {beneficiados}")

        if not beneficiados:
            app.logger.info("Nenhum dado encontrado com os filtros fornecidos.")
            return "Nenhum dado encontrado com os filtros fornecidos."

        # Criar uma nova planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Beneficiados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"

        # Adicionar a linha de cabeçalho na planilha
        ws.append([
            'Nome Beneficiado', 'CPF Beneficiado', 'Cartão SUS',
            'Nome Autorizador', 'CPF Autorizador',
            'Quantidade Liberada', 'Quantidade Restante', 'Tamanho Liberado',
            'Motivo Liberação', 'Data Início', 'Validade em Meses', 'Data Final de Validade'
        ])

        # Adicionar os dados na planilha
        for beneficiado in beneficiados:
            # Log dos dados sendo adicionados
            app.logger.info(f"Adicionando beneficiado à planilha: {beneficiado.nome_beneficiado}, CPF: {beneficiado.cpf_beneficiado}")

            autorizador = Autorizador.query.filter_by(id_usuario=beneficiado.id_autorizador).first()

            nome_autorizador = autorizador.nomeAutorizador if autorizador else 'Desconhecido'
            cpf_autorizador = autorizador.cpf if autorizador else 'Desconhecido'

            data_inicio_formatada = beneficiado.data_inicio.strftime('%d/%m/%Y') if beneficiado.data_inicio else 'Data não disponível'

            # Calcular a data final de validade
            if beneficiado.data_inicio and beneficiado.validade_meses:
                data_inicio = beneficiado.data_inicio
                validade_meses = beneficiado.validade_meses
                data_final = data_inicio + relativedelta(months=validade_meses)
                data_final_formatada = data_final.strftime('%d/%m/%Y')
            else:
                data_final_formatada = 'Data não disponível'

            ws.append([
                beneficiado.nome_beneficiado,
                beneficiado.cpf_beneficiado,
                beneficiado.cartao_sus,
                nome_autorizador,
                cpf_autorizador,
                beneficiado.quantidade_liberada,
                beneficiado.quantidade_restante,
                beneficiado.tamanho_liberado,
                beneficiado.motivo_liberacao,
                data_inicio_formatada,
                beneficiado.validade_meses,
                data_final_formatada
            ])

        # Ajustar a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Pega a letra da coluna
            for cell in col:
                if cell.value:
                    max_length = len(str(cell.value))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Salvar a planilha em um objeto BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Criar uma resposta com o arquivo Excel
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilha_beneficiados_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

        return response

    # Renderizar o template HTML se o método for GET
    return render_template('planilhaPbeneficiado.html')

@app.route('/mostrar_planilhaB', methods=['GET'])
@login_required
def mostrar_planilhaB():
    # Obter os filtros de nome, ano e mês dos parâmetros da URL
    name_filter = request.args.get('name-filter')
    ano_filter = request.args.get('ano')
    mes_filter = request.args.get('mes')

    app.logger.info(f"Filtros: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

    # Construir a consulta SQL base na tabela Beneficiado
    query = Beneficiado.query

    # Aplicar filtro pelo nome ou CPF, se fornecido
    if name_filter:
        query = query.filter(
            or_(
                Beneficiado.nome_beneficiado.ilike(f'%{name_filter}%'),
                Beneficiado.cpf_beneficiado.ilike(f'%{name_filter}%')
            )
        )

    # Aplicar filtro pelo ano, se fornecido
    if ano_filter:
        try:
            ano_filter = int(ano_filter)
            query = query.filter(extract('year', Beneficiado.data_inicio) == ano_filter)
        except ValueError:
            return "Ano inválido", 400

    # Aplicar filtro pelo mês, se fornecido
    if mes_filter:
        try:
            mes_filter = int(mes_filter)
            query = query.filter(extract('month', Beneficiado.data_inicio) == mes_filter)
        except ValueError:
            return "Mês inválido", 400

    app.logger.info(f"Consulta SQL gerada: {query}")

    # Executar a consulta e obter os resultados
    beneficiados = query.all()

    app.logger.info(f"Dados encontrados na consulta: {beneficiados}")

    if not beneficiados:
        return "Nenhum dado encontrado com os filtros fornecidos."

    # Criar uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"  # Usar 'Todos' se nenhum filtro especificado

    # Adicionar a linha de cabeçalho na planilha
    ws.append([
        'Nome Beneficiado', 'CPF Beneficiado', 'Cartão SUS', 'Quantidade Liberada',
        'Quantidade Restante', 'Tamanho Liberado', 'Motivo Liberação', 'Data Início',
        'Validade em Meses', 'Data Final de Validade'
    ])

    # Adicionar os dados na planilha
    for beneficiado in beneficiados:
        data_inicio_formatada = beneficiado.data_inicio.strftime('%Y-%m-%d') if beneficiado.data_inicio else 'Data não disponível'
        
        # Calcular a data final de validade
        if beneficiado.data_inicio and beneficiado.validade_meses:
            data_inicio = beneficiado.data_inicio
            validade_meses = beneficiado.validade_meses
            data_final = data_inicio + relativedelta(months=validade_meses)
            data_final_formatada = data_final.strftime('%Y-%m-%d')
        else:
            data_final_formatada = 'Data não disponível'
        
        ws.append([
            beneficiado.nome_beneficiado,
            beneficiado.cpf_beneficiado,
            beneficiado.cartao_sus,
            beneficiado.quantidade_liberada,
            beneficiado.quantidade_restante,
            beneficiado.tamanho_liberado,
            beneficiado.motivo_liberacao,
            data_inicio_formatada,
            beneficiado.validade_meses,
            data_final_formatada
        ])

    # Ajustar a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Pega a letra da coluna
        for cell in col:
            if cell.value:
                max_length = len(str(cell.value))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Salvar a planilha em um objeto BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Criar uma resposta com o arquivo Excel
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

    return response
# Route to generate and download the spreadsheet for Autorizador
@app.route('/gerar_planilhaA', methods=['GET', 'POST'])
def gerar_planilhaA():
    if request.method == 'POST':
        # Obter os filtros de nome, ano e mês do formulário
        name_filter = request.form.get('namee-filter')
        ano_filter = request.form.get('anoo-filter')
        mes_filter = request.form.get('mess-filter')

        app.logger.info(f"Filtros recebidos: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

        # Iniciar a query na tabela 'Autorizador' e 'TransacaoA'
        query = db.session.query(Autorizador, TransacaoA).join(TransacaoA, TransacaoA.autorizador_id == Autorizador.id)

        # Aplicar filtro pelo nome, se fornecido
        if name_filter:
            query = query.filter(
                or_(
                    Autorizador.nomeAutorizador.ilike(f'%{name_filter}%'),
                    Autorizador.cpf.ilike(f'%{name_filter}%')
                )
            )

        # Aplicar filtro pelo ano, se fornecido
        if ano_filter:
            try:
                ano_filter = int(ano_filter)
                query = query.filter(extract('year', TransacaoA.data_retirada) == ano_filter)
            except ValueError:
                return "Ano inválido", 400

        # Aplicar filtro pelo mês, se fornecido
        if mes_filter:
            try:
                mes_filter = int(mes_filter)
                query = query.filter(extract('month', TransacaoA.data_retirada) == mes_filter)
            except ValueError:
                return "Mês inválido", 400

        app.logger.info(f"Consulta SQL gerada: {query}")

        # Executar a consulta e obter os resultados
        dadosA = query.all()

        app.logger.info(f"Dados encontrados na consulta: {dadosA}")

        # Verificar se há dados encontrados na consulta
        if not dadosA:
            return "Nenhum dado encontrado com os filtros fornecidos."

        # Criar uma nova planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"  # Usar 'Todos' se nenhum filtro especificado

        # Adicionar a linha de cabeçalho na planilha
        ws.append(['Nome Fantasia', 'CNPJ', 'Quantidade', 'Marca', 'Tamanho', 'Quantidade Total', 'Media de Marca de fralda', 'Media Tamanho de fralda', 'Data'])

        # Adicionar os dados na planilha
        for autorizador, transacaoA in processar_dados_ao_final_do_mes:
            data_retirada = transacaoA.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
            ws.append([
                autorizador.nomeAutorizador,
                autorizador.cpf,
                transacaoA.quantidadeBeneficiado,
                transacaoA.marca_fralda_entregue,
                transacaoA.tamanho_fralda,
                autorizador.quantidadeTotal,
                transacaoA.media_marca_fralda,
                transacaoA.media_tamanho_fralda,
                data_retirada.strftime('%Y-%m-%d %H:%M:%S')
            ])

        # Salvar a planilha em um objeto BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Criar uma resposta com o arquivo Excel
        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

        return response

    # Renderizar o template HTML se o método for GET
    return render_template('planilhaPAutorizador.html')
@app.route('/mostrar_planilhaA', methods=['GET'])
@login_required
def mostrar_planilhaA():
    # Obter os filtros de nome, ano e mês dos parâmetros da URL
    name_filter = request.args.get('namee-filter')
    ano_filter = request.args.get('anoo')
    mes_filter = request.args.get('mess')

    app.logger.info(f"Filtros: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")
    print("1")
    # Construir a consulta SQL base
    query = db.session.query(Autorizador, TransacaoA).join(TransacaoA, Autorizador.id == TransacaoA.autorizador_id)

    # Aplicar filtro pelo nome, se fornecido
    if name_filter:
        print("2")
        query = query.filter(
            or_(
                Autorizador.nomeAutorizador.ilike(f'%{name_filter}%'),
                Autorizador.cpf.ilike(f'%{name_filter}%')
            )
        )

    # Aplicar filtro pelo ano, se fornecido
    if ano_filter:
        print("3")
        try:
            ano_filter = int(ano_filter)
            query = query.filter(extract('year', TransacaoA.data_retirada) == ano_filter)
        except ValueError:
            return "Ano inválido", 400

    # Aplicar filtro pelo mês, se fornecido
    if mes_filter:
        print("4")
        try:
            mes_filter = int(mes_filter)
            query = query.filter(extract('month', TransacaoA.data_retirada) == mes_filter)
        except ValueError:
            return "Mês inválido", 400

    app.logger.info(f"Consulta SQL gerada: {query}")

    # Executar a consulta e obter os resultados
    dadosA = query.all()

    app.logger.info(f"Dados encontrados na consulta: {dadosA}")

    if not dadosA:
        print("5")
        return "Nenhum dado encontrado com os filtros fornecidos."

    # Processar os dados e retornar a resposta adequada
    print(7)
    # Criar uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"  # Usar 'Todos' se nenhum filtro especificado

    # Adicionar a linha de cabeçalho na planilha
    ws.append(['Nome', 'CPF', 'Data do cadastro', 'Nome Beneficiado Cadastrado', 'CPF Beneficiado' , 'Total de cadastros'])

    # Adicionar os dados na planilha
    for autorizador, transacaoA in dadosA:
        print("6")
        data_retirada = transacaoA.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
        ws.append([autorizador.nomeAutorizador, autorizador.cpf,  transacaoA.data_retirada.strftime('%Y-%m-%d %H:%M:%S'), transacaoA.detalhes_beneficiados, transacaoA.cpf_beneficiadoA, transacaoA.total_por_mes])

    # Salvar a planilha em um objeto BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Criar uma resposta com o arquivo Excel
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

    return response
from flask import Flask, request, make_response, render_template
from flask_login import current_user, login_required
from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import or_, extract
import pytz

# Rota para mostrar planilha dos autorizadores
@app.route('/mostrar_planilhaAut', methods=['GET'])
@login_required
def mostrar_planilhaAut():
    ano_filter = request.args.get('ano')
    mes_filter = request.args.get('mes')
    name_filter = request.args.get('name-filter')

    app.logger.info(f"Filtros: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

    usuario = current_user
    autorizador = usuario.autorizador

    # Construir a consulta base
    query = db.session.query(Autorizador, TransacaoA).join(TransacaoA, autorizador.id == TransacaoA.autorizador_id)

    # Aplicar filtro de nome, se fornecido
    if name_filter:
        query = query.filter(
            or_(
                TransacaoA.detalhes_beneficiados.ilike(f'%{name_filter}%'),
                TransacaoA.cpf_beneficiadoA.ilike(f'%{name_filter}%')
            )
        )

    # Aplicar filtro de ano, se fornecido
    if ano_filter:
        try:
            ano_filter = int(ano_filter)
            query = query.filter(extract('year', TransacaoA.data_retirada) == ano_filter)
        except ValueError:
            return "Ano inválido", 400

    # Aplicar filtro de mês, se fornecido
    if mes_filter:
        try:
            mes_filter = int(mes_filter)
            query = query.filter(extract('month', TransacaoA.data_retirada) == mes_filter)
        except ValueError:
            return "Mês inválido", 400

    app.logger.info(f"Consulta SQL gerada: {query}")

    # Executar a consulta
    dadosA = query.all()

    app.logger.info(f"Dados encontrados na consulta: {dadosA}")

    if not dadosA:
        flash('Nenhum dado encontrado com os filtros fornecidos.')

    # Criar planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"


    # Adicionar dados na planilha
    for autorizador, transacaoA in dadosA:
        data_retirada = transacaoA.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
        ws.append([
            autorizador.nomeAutorizador,
            autorizador.cpf,
            transacaoA.data_retirada.strftime('%Y-%m-%d %H:%M:%S'),
            transacaoA.detalhes_beneficiados,
            transacaoA.cpf_beneficiadoA,
            transacaoA.total_por_mes
        ])

    # Salvar planilha em um objeto BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Criar resposta com arquivo Excel
    response = make_response(excel_file.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

    return response
# Rota para gerar planilha dos autorizadores
@app.route('/gerar_planilhaAut', methods=['GET', 'POST'])
@login_required
def gerar_planilhaAut():
    if request.method == 'POST':
        name_filter = request.form.get('name-filter')
        ano_filter = request.form.get('ano-filter')
        mes_filter = request.form.get('mes-filter')

        app.logger.info(f"Filtros recebidos: name_filter={name_filter}, ano_filter={ano_filter}, mes_filter={mes_filter}")

        # Obter o autorizador associado ao usuário autenticado
        usuario = current_user
        autorizador = usuario.autorizador

        if not autorizador:
            return "Usuário não associado a um autorizador", 400

        query = db.session.query(Autorizador, TransacaoA).join(TransacaoA, TransacaoA.autorizador_id == Autorizador.id)
        query = query.filter(Autorizador.id == autorizador.id)

        # Aplicar filtro por nome, se fornecido
        if name_filter:
            query = query.filter(
                or_(
                    TransacaoA.detalhes_beneficiados.ilike(f'%{name_filter}%'),
                    TransacaoA.cpf_beneficiadoA.ilike(f'%{name_filter}%')
                )
            )

        # Aplicar filtro de ano, se fornecido
        if ano_filter:
            try:
                ano_filter = int(ano_filter)
                query = query.filter(extract('year', TransacaoA.data_retirada) == ano_filter)
            except ValueError:
                return "Ano inválido", 400

        # Aplicar filtro de mês, se fornecido
        if mes_filter:
            try:
                mes_filter = int(mes_filter)
                query = query.filter(extract('month', TransacaoA.data_retirada) == mes_filter)
            except ValueError:
                return "Mês inválido", 400

        app.logger.info(f"Consulta SQL gerada: {query}")

        dadosA = query.all()

        app.logger.info(f"Dados encontrados na consulta: {dadosA}")

        if not dadosA:
            return "Nenhum dado encontrado com os filtros fornecidos."

        wb = Workbook()
        ws = wb.active
        ws.title = f"Dados_{mes_filter or 'Todos'}_{ano_filter or 'Todos'}"

        ws.append(['Nome', 'CPF', 'Quantidade Beneficiado', 'Total por Mês', 'Detalhes Beneficiados', 'Data'])

        for autorizador, transacaoA in dadosA:
            data_retirada = transacaoA.data_retirada.replace(tzinfo=pytz.timezone('America/Sao_Paulo'))
            ws.append([
                autorizador.nomeAutorizador,
                autorizador.cpf,
                transacaoA.quantidadeBeneficiado,
                transacaoA.total_por_mes,
                transacaoA.detalhes_beneficiados,
                data_retirada.strftime('%Y-%m-%d %H:%M:%S')
            ])

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=planilha_{mes_filter or "Todos"}_{ano_filter or "Todos"}.xlsx'

        return response

    return render_template('planilhaAutorizador.html')


# Rota para renderizar a página com informações da farmácia
@app.route('/visualizaFarmacia.html')
@login_required
def visualiza_farmacia():
    # Aqui você pode adicionar lógica para buscar as informações da farmácia do banco de dados
    # Substitua esta lógica pelo seu método real de busca de dados
    # Por exemplo:
    farmacia = {
        'nomeFantasia': 'Farmácia ABC',
        'cnpj': '12345678901234',
        'quantidadeTotal': 1000
    }
    return render_template('visuPbeneficiado.html', farmacia=farmacia)

# Rota para atualizar as informações da farmácia
@app.route('/atualizarFarmacia/<cnpj>', methods=['POST'])
@login_required
def atualizarFarmacia(cnpj):
    tipo_usuario = session.get('tipo_usuario','3')
    if tipo_usuario=='3':
        try:
            farmacia = Farmacia.query.filter_by(cnpj=cnpj).first()
            if not farmacia:
                return redirect(url_for('pagina_de_erro'))

            # Atualize as informações da farmácia com base nos dados do formulário enviado
            farmacia.nomeFantasia = request.form['nomeFantasia']
            farmacia.cnpj= request.form['cnpj']
            

 

            # Commit para salvar as mudanças no banco de dados
            flash("Dados atualizados com sucesso!")
            db.session.commit()

            # Redirecione de volta para a página de visualização da farmácia
            return '', 204
        except Exception as e:
            return render_template("prefeitura.html")

    else:
            print("entrou no else")
            return redirect(url_for('rota_protegida'))
    
@app.route('/relatorioP')
@login_required
def relatorioP():
    return render_template('relatorioP.html')

#rota de busca para farmacia
@app.route('/buscaPfarmacia')
@login_required
def farmacias():
    tipo_usuario = session['tipo_usuario']
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    else:
        nomePrefeitura = 'Prefeitura'   
    return render_template('buscaPrefeitura.html',page='buscaPFarmacia', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)
#rota de busca para beneficiado
@app.route('/buscaPbeneficiado')
@login_required
def beneficiados():
    tipo_usuario = session['tipo_usuario']
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    else:
        nomePrefeitura = 'Prefeitura'   
    return render_template('buscaPrefeitura.html', page='buscaPBeneficiado', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

#rota de busca para autorizador
@app.route('/buscaPautorizador')
@login_required
def autorizadores():
    tipo_usuario = session['tipo_usuario']
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    else:
        nomePrefeitura = 'Prefeitura'
    return render_template('buscaPrefeitura.html' , page = 'BuscaPAutorizador', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)


def registrar_log(descricao_log):
    try:
        with open('log.txt', 'a') as arquivo_log:
            arquivo_log.write(descricao_log)
    except Exception as e:
        print(f"Erro ao escrever no arquivo de log: {e}")

def get_mac_address():
    mac = uuid.UUID(int=uuid.getnode()).hex[-12:]
    return ":".join([mac[e:e+2] for e in range(0, 11, 2)])

def registrar_log_usuario(usuario, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()
    
    if hasattr(usuario, 'prefeitura') and usuario.prefeitura:
        identificador_usuario = f"CPF: {usuario.prefeitura.cpf_prefeitura}"
    elif hasattr(usuario, 'autorizador') and usuario.autorizador:
        identificador_usuario = f"CPF: {usuario.autorizador.cpf}"
    elif hasattr(usuario, 'farmacia') and usuario.farmacia:
        identificador_usuario = f"CNPJ: {usuario.farmacia.cnpj}"
    else:
        identificador_usuario = 'Identificador não disponível'
    
    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de Login:
                    - Data: {data_alteracao}
                    - ID do Usuário: {usuario.id}
                    - Email: {usuario.email}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - {identificador_usuario}
                    - IP da Máquina: {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)

def registrar_log_beneficiado(beneficiado, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()
    
    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro do beneficiado:
                    - Data: {data_alteracao}
                    - ID: {beneficiado.id}
                    - Nome: {beneficiado.nome_beneficiado}
                    - CPF: {beneficiado.cpf_beneficiado}
                    - CPF Pesquisado: {beneficiado.cpfPesquisado}
                    - Usuário Pendente: {beneficiado.usuarioPendente}
                    - Pendente: {beneficiado.pendente}
                    - Cartão SUS: {beneficiado.cartao_sus}
                    - Nome do Autorizado: {beneficiado.nome_autorizado}
                    - CPF do Autorizado: {beneficiado.cpf_autorizado}
                    - Quantidade Liberada: {beneficiado.quantidade_liberada}
                    - Quantidade Pego: {beneficiado.quantidade_pego}
                    - Quantidade Restante: {beneficiado.quantidade_restante}
                    - Quantidade Total: {beneficiado.quantidadeTotal}
                    - Tamanho Liberado: {beneficiado.tamanho_liberado}
                    - Motivo de Liberação: {beneficiado.motivo_liberacao}
                    - Data de Início: {beneficiado.data_inicio}
                    - Validade em Meses: {beneficiado.validade_meses}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina: {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)

def registrar_log_farmacia(farmacia, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()

    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro temporário de farmácia:
                    - Data: {data_alteracao}
                    - ID: {farmacia.id}
                    - CNPJ: {farmacia.cnpj}
                    - Nome Fantasia: {farmacia.nome}
                    - ID do Usuário: {farmacia.id}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina : {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)
def registrar_log_farmacia_confirmacao(farmacia, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()

    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro de farmácia:
                    - Data: {data_alteracao}
                    - ID: {farmacia.id}
                    - CNPJ: {farmacia.cnpj}
                    - Nome Fantasia: {farmacia.nomeFantasia}
                    - ID do Usuário: {farmacia.id}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina : {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)

def registrar_log_autorizador(autorizador, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()

    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro temporario de autorizador:
                    - Data: {data_alteracao}
                    - ID: {autorizador.id}
                    - CPF: {autorizador.cpf}
                    - Nome do Autorizador: {autorizador.nome}
                    - ID do Usuário: {autorizador.id}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina : {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)

def registrar_log_autorizador_confirmacao(autorizador, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()

    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro de autorizador confirmado :
                    - Data: {data_alteracao}
                    - ID: {autorizador.id}
                    - CPF: {autorizador.cpf}
                    - Nome do Autorizador: {autorizador.nomeAutorizador}
                    - ID do Usuário: {autorizador.id_usuario}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina : {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)

def registrar_log_prefeitura(prefeitura, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()

    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro temporário de prefeitura:
                    - Data: {data_alteracao}
                    - ID: {prefeitura.id}
                    - CPF Prefeitura: {prefeitura.cpf}
                    - Nome Prefeitura: {prefeitura.nome}
                    - ID do Usuário: {prefeitura.id}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina: {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)
def registrar_log_prefeitura_confirmado(prefeitura, id_usuario_logado):
    hostname = socket.gethostname()
    ip_local = socket.gethostbyname(hostname)
    mac_address = get_mac_address()

    data_alteracao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    descricao_log = f'''Registro de cadastro de prefeitura:
                    - Data: {data_alteracao}
                    - ID: {prefeitura.id}
                    - CPF Prefeitura: {prefeitura.cpf_prefeitura}
                    - Nome Prefeitura: {prefeitura.nomePrefeitura}
                    - ID do Usuário: {prefeitura.id}
                    - ID do Usuário Logado: {id_usuario_logado}
                    - IP da Máquina: {ip_local}
                    - MAC da Máquina: {mac_address}
                    \n'''
    registrar_log(descricao_log)


@app.route('/gerar_pdf_logs', methods=['POST'])
@login_required
def gerar_pdf_logs():
    try:
        
        if session.get('tipo_usuario') not in ['Farmacia', 'Autorizador', 'Prefeitura']:
            # Se não for nenhum desses, retorna uma resposta de acesso negado
            return render_template('errorPage.html')

        with open('log.txt', 'r', encoding='latin-1') as logs:
            log = logs.read()

        pdf = FPDF()
        pdf.add_page()

        # Adicione o título
        pdf.set_font("Courier", size=14)
        pdf.cell(200, 10, txt="Logs do Sistema", ln=True, align='C')

        # Adicione os logs
        pdf.set_font("Courier", size=6)
        pdf.multi_cell(0, 3, txt=log)

        temp_file_path = "logs.pdf"
        pdf.output(temp_file_path)

    except Exception as e:
        print(e)
        return render_template('errorPage.html')

    return send_file(temp_file_path, as_attachment=True)




@app.route('/busca')
@login_required
def busca():
    nomeAutorizador = ""
    nomeFantasia = ""
    tipo_usuario = session.get('tipo_usuario')
    if 'nomeAutorizador' in session:
        nomeAutorizador = session['nomeAutorizador']
    elif 'nomeFantasia' in session: 
        nomeFantasia = session ['nomeFantasia']



    return render_template("busca.html", nomeAutorizador=nomeAutorizador, nomeFantasia=nomeFantasia, tipo_usuario=tipo_usuario)

@app.route('/rota_protegida')
@login_required
def rota_protegida():
    return render_template("TelaInicial.html")
    # Se o usuário não estiver autenticado, ele será redirecionado para a página de login
    # O conteúdo da rota só será acessível se o usuário estiver autenticado
@app.route('/visualizar')
@login_required
def visualizar():
    tipo_usuario = session.get('tipo_usuario', '3')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    return render_template("visualizar.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario, page='visualizar')
@app.route('/cadastros')
def cadastros():
    tipo_usuario = session.get('tipo_usuario', '3')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    return render_template('visualizar.html', page='cadastros', nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)


@app.route('/dashboards')
@login_required
def dashboards():
    tipo_usuario = session.get('tipo_usuario', '3')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    else:
        nomePrefeitura = 'Prefeitura'

    
    return render_template("dashboards.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

@app.route('/relatorios')
@login_required
def relatorios():
    tipo_usuario = session.get('tipo_usuario', '3')
    if 'nomePrefeitura' in session:
        nomePrefeitura = session['nomePrefeitura']
    else:
        nomePrefeitura = 'Prefeitura'

    
    return render_template("relatorios.html", nomePrefeitura=nomePrefeitura, tipo_usuario=tipo_usuario)

@app.route('/cadastroAutorizador')
@login_required
def cadastroAutorizador():
    tipo_usuario = session.get('tipo_usuario', '2')
    if 'nomeAutorizador' in session:
        nomeAutorizador = session['nomeAutorizador']
    else:
        nomeAutorizador = ''
    return render_template("cadastroBenAut.html", nomeAutorizador=nomeAutorizador, tipo_usuario=tipo_usuario)

@app.route('/escolhaP')
@login_required
def visualiza():
    return render_template("escolhaP.html")


@app.route('/autorizador')
@login_required
def autorizador():
    return render_template("autorizador.html")
@app.route('/redefiDois')
@login_required
def redefDoiss():
    return render_template("redefiDois.html")
@app.route('/redef')

def redeff():
    return render_template("redef.html")

from flask import send_file, jsonify, abort
import os
import io

from flask import request, send_file, abort
import os
from datetime import datetime

from flask import send_file

from flask import send_file





@app.route('/planilhaFarmacia')
def planilhaFarmacia():
    return render_template('planilhaFarmacia.html')



@app.route("/planilhaPAutorizador", methods=['GET', 'POST'])
def planilhaPAutorizador():
    return render_template("planilhaPAutorizador.html")

def get_fralda_data():
    # Buscar dados do banco de dados usando SQLAlchemy
    transacoes = db.session.query(
        Transacao.farmacia_id,
        Farmacia.nomeFantasia,
        db.func.sum(Transacao.quantidade).label('quantidade'),
        db.func.date(Transacao.data_retirada).label('dia')
    ).join(Farmacia, Transacao.farmacia_id == Farmacia.id)\
     .group_by(Transacao.farmacia_id, db.func.date(Transacao.data_retirada))\
     .all()

    result = []
    for transacao in transacoes:
        result.append({
            'farmacia': transacao.nomeFantasia,
            'quantidade': transacao.quantidade,
            'dia': transacao.dia.strftime('%Y-%m-%d')
        })
    return result

def get_beneficiado_data():
    print("é chamado")
    # Buscar dados do banco de dados usando SQLAlchemy
    transacoes = db.session.query(
        TransacaoA.autorizador_id,
        Autorizador.nomeAutorizador,
        db.func.date(TransacaoA.data_retirada).label('dia')
    ).join(Autorizador, TransacaoA.autorizador_id== Autorizador.id)\
     .group_by(TransacaoA.autorizador_id, db.func.date(TransacaoA.data_retirada))\
     .all()

    result = []
    for transacaoA in transacoes:
        result.append({
           
            
            'dia': transacaoA.dia.strftime('%Y-%m-%d')
        })
    return result

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/dados', methods=['GET'])
def obter_dados():
    dados = get_fralda_data()
    return jsonify(dados)
@app.route('/dadosA', methods=['GET'])
def obter_dadosA():
    print("entra aqui")
    dadosA = get_beneficiado_data()
    return jsonify(dadosA)

farmacia_dash_app = create_dash_app(app)
beneficiado_dash_app = create_dashboard_beneficiado(app)
autorizador_dash_app = create_dashboard_autorizador(app)


@app.route('/dashFarmacia',  methods=['POST','GET'])
def dashFarmacia():
    return farmacia_dash_app.index()
@app.route('/dash_autorizador')
def render_dashboard_autorizador():
    return autorizador_dash_app.index()

@app.route('/dash_beneficiado', methods=['POST','GET'])
def dashBeneficiado():
    return beneficiado_dash_app.index()
@app.errorhandler(405)
def method_not_allowed(e):
    return f'Method not allowed: {request.method}', 405


@app.route('/dados_beneficiado', methods=['GET', 'POST'])
def obter_dados_beneficiado():
    beneficiados = Beneficiado.query.all()
    result = []
    for beneficiado in beneficiados:
        result.append({
            'id': beneficiado.id,
            'nome_beneficiado': beneficiado.nome_beneficiado,
            'tamanho_liberado': beneficiado.tamanho_liberado,
            'quantidade_liberada': beneficiado.quantidade_liberada,
            'id_autorizador': beneficiado.id_autorizador,
            'marca_fralda': beneficiado.marca_fralda,
            'data_inicio': beneficiado.data_inicio.strftime('%Y-%m-%d') if beneficiado.data_inicio else None
        })
    return jsonify(result)


@app.route('/dados_antigos', methods=['GET'])
def dados_antigos():
    farmacias = Farmacia.query.all()
    antigos_data = [
        {
            'farmacia': farmacia.nomeFantasia,
            'quantidade total': farmacia.quantidade,
            'tipo': 'total',  # Adiciona um tipo para diferenciar os dados antigos dos novos
            'data': farmacia.data_retirada.strftime('%Y-%m-%d') if farmacia.data_retirada else None
        }
        for farmacia in farmacias
    ]
    return jsonify(antigos_data)

@app.route('/dados_transacao', methods=['GET'])
def dados_transacao():
    transacoes = Transacao.query.all()
    # media_tamanho_fralda = media_tamanho_fralda.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)

    transacoes_data = [
        {
            'dia': transacao.data_retirada.strftime('%Y-%m-%d'),
            'farmacia': transacao.farmacia.nomeFantasia,
            'quantidade': transacao.quantidade, 
            'marca': transacao.marca_fralda_entregue,
            'tamanho': transacao.tamanho_fralda,
            #  'media tamanho mais retirado': transacao.media_tamanho_fralda.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP),
            # 'media marca mais retirada': transacao.media_marca_fralda.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP),
            'tipo': 'diario'  # Adiciona um tipo para diferenciar os dados diários dos totais
        }
        for transacao in transacoes
    ]
    return jsonify(transacoes_data)

@app.route('/dados_farmacia', methods=['GET'])
def dados_farmacia():
    farmacias = Farmacia.query.all()
    farmacias_data = [
        {
            'nomeFantasia': farmacia.nomeFantasia,
            'bairro': farmacia.bairro   ,
            'cidade': farmacia.cidade,
            'estado': farmacia.estado,
            'cep': farmacia.cep,
            'numero': farmacia.numero,
            'rua': farmacia.logradouro,
            'total_fraldas_entregues': farmacia.quantidade
        }
        for farmacia in farmacias
    ]
    return jsonify(farmacias_data)
@app.route('/dados_transacaoA', methods=['GET'])
def dados_transacaoA():
    transacoes = TransacaoA.query.all()
    transacoes_data = [
        {
            'dia': transacaoA.data_retirada.strftime('%Y-%m-%d'),
            'quantidade': transacaoA.quantidadeBeneficiado,
            'tipo': 'diario'  # Adiciona um tipo para diferenciar os dados diários dos totais
        }
        for transacaoA in transacoes
    ]
    return jsonify(transacoes_data)


@app.route('/dados_autorizador', methods=['GET'])
def dados_autorizador():
    autorizadores = Autorizador.query.all()
    result = [{'id': aut.id, 'nomeAutorizador': aut.nomeAutorizador, 'id_usuario': aut.id_usuario} for aut in autorizadores]
    return jsonify(result)


def media_tamanho():
    resultados = db.session.query(
        func.strftime('%Y-%m', Transacao.data_retirada).label('mes'),
        Transacao.farmacia_id,
        Transacao.tamanho_fralda
    ).all()

    # Converter resultados para DataFrame
    df = pd.DataFrame(resultados, columns=['mes', 'id_farmacia', 'tamanho_fralda'])

    # Calcular a moda do tamanho de fraldas por mês e farmácia
    moda_tamanho_mes = df.groupby(['mes', 'id_farmacia'])['tamanho_fralda'].agg(lambda x: x.mode().iloc[0])

    # Salvar ou atualizar os resultados na tabela Dados
    for (mes, id_farmacia), moda_tamanho in moda_tamanho_mes.items():
        registro_existente = db.session.query(Dados).filter_by(
            id_farmacia=id_farmacia,
            mes=mes
        ).first()

        if registro_existente:
            registro_existente.media_tamanho_fralda = moda_tamanho
        else:
            novo_dado = Dados(
                id_farmacia=id_farmacia,
                mes=mes,
                media_tamanho_fralda=moda_tamanho
            )
            db.session.add(novo_dado)

    db.session.commit()

    return moda_tamanho_mes

@app.route('/calcular_media_tamanho', methods=['GET'])
def calcular_media_tamanho():
    media_tamanho()
    return "Moda do tamanho de fraldas calculada e salva no banco de dados."


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'documentacao' not in request.files or 'usuario_temp_id' not in request.form:
        logging.error("Arquivo ou ID do usuário temporário não encontrado no request.")
        return redirect(request.url)
    
    file = request.files['documentacao']
    usuario_temp_id = request.form['usuario_temp_id']
    
    if file.filename == '':
        logging.error("Nenhum arquivo foi selecionado.")
        return redirect(request.url)
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        if filename.rsplit('.', 1)[1].lower() != 'pdf':
            pdf_filename = filename.rsplit('.', 1)[0] + '.pdf'
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
            convert_image_to_pdf(file_path, pdf_path)
            os.remove(file_path) 
            file_path = pdf_path

        with open(file_path, 'rb') as f:
            file_data = f.read()

        usuario_temp = UsuarioTemp.query.get(usuario_temp_id)
        if usuario_temp:
            usuario_temp.documento_liberacao = file_data  
            try:
                db.session.commit()
                logging.info("Documento salvo com sucesso no banco de dados.")
            except Exception as e:
                logging.error(f"Erro ao salvar no banco de dados: {e}")
                db.session.rollback()
        else:
            logging.error("Usuário temporário não encontrado.")
        
        return redirect(url_for('upload_success'))

    return 'Arquivo não permitido'
@app.route('/upload_success')
def upload_success():
    return "Upload bem-sucedido!"

@app.route('/desativar', methods=['POST'])
def desativar():
    cpf_pesquisado = session.get('cpf_pesquisado')
    autorizador = Autorizador.query.filter_by(cpf=cpf_pesquisado).first()  # Aqui você deve ajustar a lógica para obter o autorizador correto
    if autorizador:
        user = Usuario.query.filter_by(id=autorizador.id_usuario).first()
        if user:
            user.ativo = False
            db.session.commit()
            flash('Usuário desativado com sucesso.', 'success')
        else:
            flash('Usuário não encontrado.', 'error')
    else:
        flash('Autorizador não encontrado.', 'error')

    return '', 204

@app.route('/ativar', methods=['POST'])
def ativar():
    cpf_pesquisado = session.get('cpf_pesquisado')
    autorizador = Autorizador.query.filter_by(cpf=cpf_pesquisado).first() # Aqui você deve ajustar a lógica para obter o autorizador correto
    if autorizador:
        user = Usuario.query.filter_by(id=autorizador.id_usuario).first()
        if user:
            print(user.id)
            user.ativo = True
            db.session.commit()
            flash('Usuário ativado com sucesso.', 'success')
        else:
            flash('Usuário não encontrado.', 'error')
    else:
        flash('Autorizador não encontrado.', 'error')

    return '', 204
@app.route('/desativar_farmacia', methods=['POST'])
def desativar_farmacia():
    cnpj_pesquisado = session.get('cnpj_pesquisado')
    farmacia =Farmacia.query.filter_by(cnpj=cnpj_pesquisado).first()  # Aqui você deve ajustar a lógica para obter o autorizador correto
    if farmacia:
        user = Usuario.query.filter_by(id=farmacia.id_usuario).first()
        if user:
            user.ativo = False
            db.session.commit()
            flash('Usuário desativado com sucesso.', 'success')
        else:
            flash('Usuário não encontrado.', 'error')
    else:
        flash('Farmacia não encontrado.', 'error')

    return '', 204

@app.route('/ativar_farmacia', methods=['POST'])
def ativar_farmacia():
    cnpj_pesquisado = session.get('cnpj_pesquisado')
    farmacia =Farmacia.query.filter_by(cnpj=cnpj_pesquisado).first()  # Aqui você deve ajustar a lógica para obter o autorizador correto
    if farmacia:
        user = Usuario.query.filter_by(id=farmacia.id_usuario).first()
        if user:
            print(user.id)
            user.ativo = True
            db.session.commit()
            flash('Usuário ativado com sucesso.', 'success')
        else:
            flash('Usuário não encontrado.', 'error')
    else:
        flash('Autorizador não encontrado.', 'error')

    return '', 204
@app.route('/desativar_prefeitura', methods=['POST'])
def desativar_prefeitura():
    cnpj_pesquisado = session.get('cnpj_pesquisado')
    farmacia =Farmacia.query.filter_by(cnpj=cnpj_pesquisado).first()  # Aqui você deve ajustar a lógica para obter o autorizador correto
    if farmacia:
        user = Usuario.query.filter_by(id=farmacia.id_usuario).first()
        if user:
            user.ativo = False
            db.session.commit()
            flash('Usuário desativado com sucesso.', 'success')
        else:
            flash('Usuário não encontrado.', 'error')
    else:
        flash('Autorizador não encontrado.', 'error')

    return '', 204

@app.route('/ativar_prefeitura', methods=['POST'])
def ativar_prefeitura():
    cpf_pesquisado = session.get('cpf_pesquisado')
    autorizador = Autorizador.query.filter_by(cpf=cpf_pesquisado).first() # Aqui você deve ajustar a lógica para obter o autorizador correto
    if autorizador:
        user = Usuario.query.filter_by(id=autorizador.id_usuario).first()
        if user:
            print(user.id)
            user.ativo = True
            db.session.commit()
            flash('Usuário ativado com sucesso.', 'success')
        else:
            flash('Usuário não encontrado.', 'error')
    else:
        flash('Autorizador não encontrado.', 'error')
    return '', 204

@app.route('/list-files-farmacia/<int:farmacia_id>', methods=['GET'])    
def list_files_farmacia(farmacia_id):
    app.logger.info(f"Recebendo solicitação para listar arquivos da farmácia com ID: {farmacia_id}")
    farmacia = Farmacia.query.filter_by(id=farmacia_id).first()
    if farmacia:
        documentos = []
        app.logger.info(f"Farmácia encontrada: {farmacia.id}")
        if farmacia.documento_liberacao:
            app.logger.info(f"Documento encontrado para farmácia {farmacia.id}")
            documentos.append({
                'filename': f"{farmacia.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf",
                'url': url_for('get_documents_farmacia', farmacia_id=farmacia_id)
            })
        else:
            app.logger.info(f"Nenhum documento para farmácia {farmacia.id}")
        return jsonify(documentos)
    
    app.logger.info(f"Farmácia não encontrada com ID: {farmacia_id}")
    return jsonify([]), 404

@app.route('/get-documento-farmacia/<int:farmacia_id>', methods=['GET'])
@login_required
def get_documents_farmacia(farmacia_id):
    
    farmacia = Farmacia.query.filter_by(id=farmacia_id).first()
    if farmacia and farmacia.documento_liberacao:
        documento_blob = farmacia.documento_liberacao
        file_name = f"{farmacia_id}_documento_{datetime.now().strftime('%Y-%m-%d')}.pdf"
        return send_file(io.BytesIO(documento_blob), mimetype='application/pdf', as_attachment=True, download_name=file_name)
    else:
        
        return "Documento não encontrado!", 404
    
@app.route('/get-documento-autorizador/<int:autorizador_id>', methods=['GET'])
@login_required
def get_documento_autorizador(autorizador_id):
    autorizador = Autorizador.query.filter_by(id=autorizador_id).first()
    if autorizador and autorizador.documento_liberacao:
        documento_blob = autorizador.documento_liberacao
        file_name = f"{autorizador_id}_documento_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
        return send_file(
            io.BytesIO(documento_blob),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=file_name
        )
    else:
        return "Arquivo não encontrado", 404

@app.route('/list-files-autorizador/<int:autorizador_id>', methods=['GET'])
def list_files_autorizado(autorizador_id):
    autorizador = Autorizador.query.filter_by(id=autorizador_id).first()
    if autorizador:
        documentos = []
        app.logger.info(f"Autorizador encontrado: {autorizador.id}")
        if autorizador.documento_liberacao:
            app.logger.info(f"Documento encontrado para autorizador {autorizador.id}")
            documentos.append({
                'filename': f"{autorizador.cpf}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf",
                'url': url_for('get_documento_autorizador', autorizador_id=autorizador_id)
            })
        else:
            app.logger.info(f"Nenhum documento para autorizador {autorizador.id}")
        return jsonify(documentos)
    app.logger.info(f"Autorizador não encontrado: {autorizador_id}")
    return jsonify([])


@app.route('/editar_paginainicial', methods=['GET', 'POST'])
@login_required
def editar_paginainicial():
    pagina = PaginaInicial.query.first()
    
    if not pagina:
        flash('Nenhuma informação disponível para edição.')
        return redirect(url_for('inicio_route.index'))  # Redirecione para a página inicial
    
    if request.method == 'POST':
        pagina.titulo = request.form['titulo']
        pagina.descricao = request.form['descricao']
        pagina.beneficiarios_adultos = request.form['beneficiarios_adultos']
        pagina.beneficiarios_criancas = request.form['beneficiarios_criancas']
        pagina.objetivo1 = request.form['objetivo1']
        pagina.objetivo2 = request.form['objetivo2']
        pagina.objetivo3 = request.form['objetivo3']
        
        if 'capa' in request.files:
            file = request.files['capa']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                pagina.capa = filename  # Armazena apenas o nome do arquivo
            else:
                flash('Somente arquivos .jpg, .png e .tiff são permitidos.')
        
        db.session.commit()
        flash('Informações atualizadas com sucesso!')
        return redirect(url_for('editar_paginainicial'))
    
    return render_template('editar_paginainicial.html', 
                           titulo=pagina.titulo,
                           descricao=pagina.descricao,
                           beneficiarios_adultos=pagina.beneficiarios_adultos,
                           beneficiarios_criancas=pagina.beneficiarios_criancas,
                           objetivo1=pagina.objetivo1,
                           objetivo2=pagina.objetivo2,
                           objetivo3=pagina.objetivo3,
                           capa=pagina.capa)


def create_root_user():
    root_user = Usuario.query.filter_by(username='root').first()
    if not root_user:
        # Aqui use um hash seguro para a senha
        root_user = Usuario(username='root', email='root@example.com', password='hashed_password', is_admin=True, tipo_usuario=3,is_root=True)
        db.session.add(root_user)
        db.session.commit()
        print("Root user created with admin privileges and user_type 3")
    else:
        print("Root user already exists")

def setup_database(app):
    with app.app_context():
        db.create_all()
        create_root_user()


socketio.init_app(app)
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if PaginaInicial.query.first() is None:
            pagina = PaginaInicial(
                titulo="Título Padrão",
                descricao="Descrição Padrão",
                beneficiarios_adultos="Beneficiários Adultos Padrão",
                beneficiarios_criancas="Beneficiários Crianças Padrão",
                objetivo1="Objetivo 1 Padrão",
                objetivo2="Objetivo 2 Padrão",
                objetivo3="Objetivo 3 Padrão",
                capa=None
            )
            db.session.add(pagina)
            db.session.commit()

    app.run(debug=True)
    # app.run(host='0.0.0.0',port=5000)
